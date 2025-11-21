import os
import re
import uuid
import logging
import requests
from datetime import datetime, date, time, timedelta
from flask import Flask, render_template, request, redirect, url_for, jsonify, session, flash
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import psycopg2
from psycopg2 import pool
import traceback
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import secrets
import sys
from pathlib import Path
from functools import lru_cache
from flask import send_file
from io import BytesIO, StringIO
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask_socketio import SocketIO, emit, join_room, leave_room
from datetime import datetime
import json

# ============================================
#  STEP 1: Find and Load .env File
# ============================================
BASE_DIR = Path(__file__).resolve().parent
ENV_PATH = BASE_DIR / '.env'

env_loaded = False

# Try loading .env from multiple locations
if ENV_PATH.exists():
    load_dotenv(ENV_PATH)
    env_loaded = True
    print(f"✅ Loaded .env from: {ENV_PATH}")
else:
    PARENT_ENV = BASE_DIR.parent / '.env'
    if PARENT_ENV.exists():
        load_dotenv(PARENT_ENV)
        env_loaded = True
        print(f"✅ Loaded .env from: {PARENT_ENV}")
    else:
        CWD_ENV = Path.cwd() / '.env'
        if CWD_ENV.exists():
            load_dotenv(CWD_ENV)
            env_loaded = True
            print(f"✅ Loaded .env from: {CWD_ENV}")

if not env_loaded:
    print(f"⚠️  WARNING: .env file not found!")
    print(f"   Searched in:")
    print(f"   1. {ENV_PATH}")
    print(f"   2. {BASE_DIR.parent / '.env'}")
    print(f"   3. {Path.cwd() / '.env'}")
    print(f"\n   Please create .env file in one of these locations.")

# ============================================
#  STEP 2: Create Flask App (BEFORE SocketIO!)
# ============================================
app = Flask(__name__)

# ============================================
#  STEP 3: Configure Secret Key
# ============================================
SECRET_KEY = os.getenv('SECRET_KEY')
if not SECRET_KEY:
    print("ERROR: SECRET_KEY must be set in production!")
    sys.exit(1)

app.secret_key = SECRET_KEY
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
app.config['JSON_SORT_KEYS'] = False
app.config['SESSION_COOKIE_SECURE'] = False
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

print(f"✅ Secret key configured: {SECRET_KEY[:16]}...")

# ============================================
#  STEP 4: Initialize SocketIO (AFTER app creation)
# ============================================
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# Track active users
active_users = {}
active_connections = 0

# ============================================
#  STEP 5: Other Flask Configurations
# ============================================
CORS(app, origins=[
    "http://localhost:5000", 
    "http://localhost:3000",
    "https://*.ngrok.io",  # Allow ngrok
    "http://YOUR-EC2-IP",  # Replace with your EC2 public IP
    "*"  # Allow all (for testing only)
])

limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

# ============================================
#  Debug Information
# ============================================
print("=" * 60)
print("🔍 Environment Debug Information")
print("=" * 60)
print(f"Current Working Directory: {Path.cwd()}")
print(f"Script Directory: {BASE_DIR}")
print(f"Looking for .env at: {ENV_PATH}")
print(f".env exists: {ENV_PATH.exists()}")
print(f"SECRET_KEY loaded: {'Yes' if os.getenv('SECRET_KEY') else 'No'}")
print(f"RASA_DB_HOST: {os.getenv('RASA_DB_HOST', 'Not set')}")
print(f"JOBLIST_DB_HOST: {os.getenv('JOBLIST_DB_HOST', 'Not set')}")
print("=" * 60)

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s [%(levelname)s] %(message)s'
)
logger = logging.getLogger(__name__)

# Environment variables
RASA_URL = os.getenv("RASA_URL", "http://localhost:5005")
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY", "")

# Database connection pools - Initialize as None first
db_pool = None
joblist_db_pool = None

def initialize_database_pools():
    """Initialize database connection pools with proper error handling"""
    global db_pool, joblist_db_pool
    
    # Initialize rasa_db pool
    try:
        # Check if environment variables are set
        required_vars = {
            "RASA_DB_USER": os.getenv("RASA_DB_USER"),
            "RASA_DB_PASSWORD": os.getenv("RASA_DB_PASSWORD"),
            "RASA_DB_HOST": os.getenv("RASA_DB_HOST"),
            "RASA_DB_NAME": os.getenv("RASA_DB_NAME")
        }
        
        missing_vars = [k for k, v in required_vars.items() if not v]
        
        if missing_vars:
            logger.error(f"Missing required environment variables for rasa_db: {', '.join(missing_vars)}")
            logger.error("Please check your .env file")
            raise Exception(f"Missing environment variables: {', '.join(missing_vars)}")
        
        logger.info(f"Connecting to rasa_db:")
        logger.info(f"  Host: {required_vars['RASA_DB_HOST']}")
        logger.info(f"  Port: {os.getenv('RASA_DB_PORT', '5432')}")
        logger.info(f"  Database: {required_vars['RASA_DB_NAME']}")
        logger.info(f"  User: {required_vars['RASA_DB_USER']}")
        
        db_pool = psycopg2.pool.SimpleConnectionPool(
            1, 20,
            user=required_vars["RASA_DB_USER"],
            password=required_vars["RASA_DB_PASSWORD"],
            host=required_vars["RASA_DB_HOST"],
            port=os.getenv("RASA_DB_PORT", "5432"),
            database=required_vars["RASA_DB_NAME"]
        )
        
        if db_pool is None:
            raise Exception("Failed to initialize rasa_db connection pool")
        
        # Test the connection
        test_conn = db_pool.getconn()
        test_conn.close()
        db_pool.putconn(test_conn)
        
        logger.info("✓ rasa_db connection pool initialized successfully")
        
    except Exception as e:
        logger.error(f"✗ Error initializing rasa_db connection pool: {e}")
        logger.error("\nTroubleshooting steps:")
        logger.error("1. Verify PostgreSQL is running")
        logger.error("2. Check database 'rasa_db' exists")
        logger.error("3. Verify username and password in .env")
        logger.error("4. Ensure .env file is in the correct directory")
        raise

    # Initialize joblist_db pool
    try:
        required_vars_joblist = {
            "JOBLIST_DB_USER": os.getenv("JOBLIST_DB_USER"),
            "JOBLIST_DB_PASSWORD": os.getenv("JOBLIST_DB_PASSWORD"),
            "JOBLIST_DB_HOST": os.getenv("JOBLIST_DB_HOST"),
            "JOBLIST_DB_NAME": os.getenv("JOBLIST_DB_NAME")
        }
        
        logger.info(f"Connecting to joblist_db:")
        logger.info(f"  Host: {required_vars_joblist['JOBLIST_DB_HOST']}")
        logger.info(f"  Port: {os.getenv('JOBLIST_DB_PORT', '5432')}")
        logger.info(f"  Database: {required_vars_joblist['JOBLIST_DB_NAME']}")
        logger.info(f"  User: {required_vars_joblist['JOBLIST_DB_USER']}")
        
        joblist_db_pool = psycopg2.pool.SimpleConnectionPool(
            1, 20,
            user=required_vars_joblist["JOBLIST_DB_USER"],
            password=required_vars_joblist["JOBLIST_DB_PASSWORD"],
            host=required_vars_joblist["JOBLIST_DB_HOST"],
            port=os.getenv("JOBLIST_DB_PORT", "5432"),
            database=required_vars_joblist["JOBLIST_DB_NAME"]
        )
        
        if joblist_db_pool is None:
            raise Exception("Failed to initialize joblist_db connection pool")
        
        # Test the connection
        test_conn = joblist_db_pool.getconn()
        test_conn.close()
        joblist_db_pool.putconn(test_conn)
        
        logger.info("✓ joblist_db connection pool initialized successfully")
        
    except Exception as e:
        logger.error(f"✗ Error initializing joblist_db connection pool: {e}")
        logger.warning("Joblist features will be disabled")
        joblist_db_pool = None

# Initialize pools
try:
    initialize_database_pools()
except Exception as e:
    logger.critical(f"Failed to initialize database pools: {e}")
    logger.critical("Application cannot start without database connection")
    exit(1)

# Database configurations (legacy, for backward compatibility)
DB_CONFIG = {
    "host": os.getenv("DB_HOST", "localhost"),
    "port": os.getenv("DB_PORT", "5432"),
    "database": os.getenv("DB_NAME", "ileco1_db"),
    "user": os.getenv("DB_USER", "postgres"),
    "password": os.getenv("DB_PASSWORD", "postgres")
}
DB_JOBLIST = {
    "host": os.getenv("JOBLIST_HOST", "localhost"),
    "port": os.getenv("JOBLIST_PORT", "5432"),
    "database": os.getenv("JOBLIST_NAME", "joblist_db"),
    "user": os.getenv("JOBLIST_USER", "postgres"),
    "password": os.getenv("JOBLIST_PASSWORD", "postgres")
}

logger.info(f"Main DB Config: {DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}")
logger.info(f"Joblist DB Config: {DB_JOBLIST['host']}:{DB_JOBLIST['port']}/{DB_JOBLIST['database']}")

def get_rasa_connection():
    """Get a connection from the rasa_db pool"""
    global db_pool
    if db_pool is None:
        logger.error("Database pool is not initialized")
        return None
    try:
        conn = db_pool.getconn()
        logger.debug("Successfully acquired connection from rasa_db pool")
        return conn
    except Exception as e:
        logger.error(f"Error getting connection from rasa_db pool: {e}")
        return None

def get_joblist_connection():
    """Get a connection from the joblist_db pool"""
    global joblist_db_pool
    if joblist_db_pool is None:
        logger.error("Joblist database pool is not initialized")
        return None
    try:
        conn = joblist_db_pool.getconn()
        logger.debug("Successfully acquired connection from joblist_db pool")
        return conn
    except Exception as e:
        logger.error(f"Error getting connection from joblist_db pool: {e}")
        return None

def release_db_connection(conn, pool_obj):
    """Release a connection back to the specified pool"""
    if conn and pool_obj:
        try:
            pool_obj.putconn(conn)
            logger.debug(f"Connection released to pool")
        except Exception as e:
            logger.error(f"Error releasing connection to pool: {e}")
            try:
                conn.close()
            except:
                pass

def close_pools():
    """Close all database connection pools"""
    global db_pool, joblist_db_pool
    try:
        if db_pool:
            db_pool.closeall()
            logger.info("rasa_db connection pool closed")
        if joblist_db_pool:
            joblist_db_pool.closeall()
            logger.info("joblist_db connection pool closed")
    except Exception as e:
        logger.error(f"Error closing connection pools: {e}")

import atexit
atexit.register(close_pools)

@socketio.on('connect')
def handle_connect():
    """Handle new WebSocket connection"""
    global active_connections
    active_connections += 1
    
    session_id = request.sid
    user_id = session.get('user_id')
    username = session.get('username', 'Anonymous')
    
    active_users[session_id] = {
        'user_id': user_id,
        'username': username,
        'connected_at': datetime.now().isoformat(),
        'role': session.get('role', 'viewer')
    }
    
    logger.info(f"✅ WebSocket connected: {username} (SID: {session_id})")
    
    # Send welcome message with current stats
    emit('connection_success', {
        'message': 'Connected to real-time updates',
        'username': username,
        'active_users': active_connections,
        'timestamp': datetime.now().isoformat()
    })
    
    # Broadcast user count update to all clients
    socketio.emit('user_count_update', {
        'active_users': active_connections,
        'timestamp': datetime.now().isoformat()
    })
    
    # Send initial dashboard stats
    send_dashboard_stats()

@socketio.on('disconnect')
def handle_disconnect():
    """Handle WebSocket disconnection"""
    global active_connections
    active_connections = max(0, active_connections - 1)
    
    session_id = request.sid
    user_info = active_users.pop(session_id, {})
    username = user_info.get('username', 'Unknown')
    
    logger.info(f"👋 WebSocket disconnected: {username} (SID: {session_id})")
    
    # Broadcast user count update
    socketio.emit('user_count_update', {
        'active_users': active_connections,
        'timestamp': datetime.now().isoformat()
    })

@socketio.on('request_stats')
def handle_stats_request():
    """Handle manual stats refresh request"""
    send_dashboard_stats()

@socketio.on('ping')
def handle_ping():
    """Handle ping to keep connection alive"""
    emit('pong', {'timestamp': datetime.now().isoformat()})

# ============================================
#  STEP 4: ADD BROADCAST HELPER FUNCTIONS
#  (After the socketio event handlers)
# ============================================

def send_dashboard_stats():
    """Send current dashboard statistics to client"""
    try:
        complaints = get_unified_complaints()
        queue = get_agent_queue()
        
        today_start = datetime.combine(date.today(), time(0, 0))
        today_end = datetime.combine(date.today(), time(23, 59, 59))
        
        total_active = len(complaints)
        in_queue = len(queue)
        critical_count = sum(1 for c in complaints if c.get('priority', '').upper() == 'CRITICAL')
        resolved_today = sum(1 for c in complaints 
                            if c['status'] == 'RESOLVED' and 
                            c['full_data'].get('timestamp') and 
                            today_start <= c['full_data']['timestamp'] <= today_end)
        
        stats = {
            'total_active': total_active,
            'in_queue': in_queue,
            'critical_count': critical_count,
            'resolved_today': resolved_today,
            'timestamp': datetime.now().isoformat()
        }
        
        emit('stats_update', stats)
        return stats
        
    except Exception as e:
        logger.error(f"Error sending dashboard stats: {e}")
        return None

def broadcast_new_complaint(complaint_data):
    """Broadcast new complaint notification to all connected clients"""
    try:
        notification = {
            'type': 'new_complaint',
            'complaint': {
                'id': complaint_data.get('record_id'),
                'customer': complaint_data.get('customer_name'),
                'issue_type': complaint_data.get('issue_type'),
                'priority': complaint_data.get('priority'),
                'timestamp': datetime.now().isoformat()
            },
            'message': f"New {complaint_data.get('priority', 'HIGH')} priority complaint received"
        }
        
        socketio.emit('new_complaint', notification)
        logger.info(f"📢 Broadcasted new complaint: {complaint_data.get('record_id')}")
        
        # Also update stats
        stats = get_current_stats()
        if stats:
            socketio.emit('stats_update', stats)
        
    except Exception as e:
        logger.error(f"Error broadcasting new complaint: {e}")

def broadcast_status_update(table, record_id, old_status, new_status):
    """Broadcast status change to all clients"""
    try:
        update = {
            'type': 'status_change',
            'table': table,
            'record_id': record_id,
            'old_status': old_status,
            'new_status': new_status,
            'timestamp': datetime.now().isoformat()
        }
        
        socketio.emit('status_update', update)
        logger.info(f"📢 Broadcasted status update: {table}/{record_id} -> {new_status}")
        
        # Update stats
        stats = get_current_stats()
        if stats:
            socketio.emit('stats_update', stats)
        
    except Exception as e:
        logger.error(f"Error broadcasting status update: {e}")

def broadcast_critical_alert(complaint_data):
    """Broadcast critical priority alert"""
    try:
        alert = {
            'type': 'critical_alert',
            'complaint': complaint_data,
            'message': '🚨 CRITICAL PRIORITY COMPLAINT',
            'timestamp': datetime.now().isoformat(),
            'sound': True
        }
        
        socketio.emit('critical_alert', alert)
        logger.warning(f"🚨 CRITICAL ALERT broadcasted: {complaint_data.get('record_id')}")
        
    except Exception as e:
        logger.error(f"Error broadcasting critical alert: {e}")

def get_current_stats():
    """Get current statistics for broadcasting"""
    try:
        complaints = get_unified_complaints()
        queue = get_agent_queue()
        
        today_start = datetime.combine(date.today(), time(0, 0))
        today_end = datetime.combine(date.today(), time(23, 59, 59))
        
        return {
            'total_active': len(complaints),
            'in_queue': len(queue),
            'critical_count': sum(1 for c in complaints if c.get('priority', '').upper() == 'CRITICAL'),
            'resolved_today': sum(1 for c in complaints 
                                if c['status'] == 'RESOLVED' and 
                                c['full_data'].get('timestamp') and 
                                today_start <= c['full_data']['timestamp'] <= today_end),
            'timestamp': datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Error getting current stats: {e}")
        return {}

def create_users_table():
    """Create users table for authentication with improved error handling"""
    print("=" * 60)
    print("🔧 Creating Users Table and Authentication System")
    print("=" * 60)
    
    conn = get_rasa_connection()
    if not conn:
        logger.error("❌ Failed to connect to database for users table creation")
        return False
    
    try:
        cur = conn.cursor()
        
        # Drop existing tables if they exist (start fresh)
        print("🗑️  Dropping existing tables if they exist...")
        cur.execute("DROP TABLE IF EXISTS audit_log CASCADE")
        cur.execute("DROP TABLE IF EXISTS user_sessions CASCADE")
        cur.execute("DROP TABLE IF EXISTS users CASCADE")
        conn.commit()
        print("✅ Existing tables dropped")
        
        # Create users table
        print("📋 Creating users table...")
        cur.execute("""
            CREATE TABLE users (
                id SERIAL PRIMARY KEY,
                username VARCHAR(100) UNIQUE NOT NULL,
                password_hash VARCHAR(255) NOT NULL,
                email VARCHAR(255) UNIQUE,
                full_name VARCHAR(255),
                role VARCHAR(50) DEFAULT 'operator',
                is_active BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                last_login TIMESTAMP,
                failed_login_attempts INTEGER DEFAULT 0,
                locked_until TIMESTAMP
            )
        """)
        conn.commit()
        print("✅ Users table created")
        
        # Verify table was created
        cur.execute("""
            SELECT column_name, data_type 
            FROM information_schema.columns 
            WHERE table_name = 'users'
            ORDER BY ordinal_position
        """)
        columns = cur.fetchall()
        print(f"✅ Verified {len(columns)} columns in users table:")
        for col in columns:
            print(f"   - {col[0]} ({col[1]})")
        
        # Create default admin user
        print("👤 Creating default admin user...")
        default_password = generate_password_hash('admin123')
        print(f"   Password hash: {default_password[:50]}...")
        
        cur.execute("""
            INSERT INTO users (username, password_hash, email, full_name, role)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (username) DO NOTHING
            RETURNING id
        """, ('admin', default_password, 'admin@ileco1.com', 'System Administrator', 'admin'))
        
        result = cur.fetchone()
        if result:
            print(f"✅ Admin user created with ID: {result[0]}")
        else:
            print("ℹ️  Admin user already exists")
        
        conn.commit()
        
        # Create sessions table
        print("📋 Creating user_sessions table...")
        cur.execute("""
            CREATE TABLE user_sessions (
                id SERIAL PRIMARY KEY,
                user_id INTEGER REFERENCES users(id),
                session_token VARCHAR(255) UNIQUE,
                ip_address VARCHAR(45),
                user_agent TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at TIMESTAMP,
                is_active BOOLEAN DEFAULT TRUE
            )
        """)
        conn.commit()
        print("✅ User sessions table created")
        
        # Create audit log table
        print("📋 Creating audit_log table...")
        cur.execute("""
            CREATE TABLE audit_log (
                id SERIAL PRIMARY KEY,
                user_id INTEGER REFERENCES users(id),
                action VARCHAR(100),
                details TEXT,
                ip_address VARCHAR(45),
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()
        print("✅ Audit log table created")
        
        # Final verification
        print("🔍 Verifying all tables...")
        cur.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_name IN ('users', 'user_sessions', 'audit_log')
            ORDER BY table_name
        """)
        tables = cur.fetchall()
        print(f"✅ Found {len(tables)} authentication tables:")
        for table in tables:
            print(f"   - {table[0]}")
        
        # Verify admin user
        cur.execute("SELECT username, email, role FROM users WHERE username = 'admin'")
        admin = cur.fetchone()
        if admin:
            print(f"✅ Admin user verified: {admin[0]} ({admin[1]}) - Role: {admin[2]}")
        
        print("=" * 60)
        print("✅ Users table and authentication system created successfully")
        print("=" * 60)
        print("📋 Default Credentials:")
        print("   Username: admin")
        print("   Password: admin123")
        print("=" * 60)
        
        return True
        
    except Exception as e:
        logger.error(f"❌ Error creating users table: {e}")
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        conn.rollback()
        return False
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

# Call this function on startup
print("\n🚀 Initializing authentication system...")
if create_users_table():
    print("✅ Authentication system ready!")
else:
    print("❌ Failed to initialize authentication system!")
    print("⚠️  The application may not work correctly.")

@lru_cache(maxsize=128)
def get_cached_complaints(cache_key):
    # Cache for 30 seconds
    return get_unified_complaints()

def get_cache_key():
    return f"{datetime.now().strftime('%Y%m%d%H%M%S')[:-1]}"  # Round to 10s


# ============================================
#  AUTHENTICATION DECORATORS
# ============================================

def login_required(f):
    """Decorator to require login for routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login to access this page', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Decorator to require admin role"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login to access this page', 'error')
            return redirect(url_for('login'))
        
        if session.get('role') != 'admin':
            flash('Admin access required', 'error')
            return redirect(url_for('dashboard'))
        
        return f(*args, **kwargs)
    return decorated_function

# ============================================
#  AUTHENTICATION HELPER FUNCTIONS
# ============================================

def log_audit(user_id, action, details, ip_address):
    """Log user actions for security audit"""
    conn = get_rasa_connection()
    if not conn:
        return
    
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO audit_log (user_id, action, details, ip_address)
            VALUES (%s, %s, %s, %s)
        """, (user_id, action, details, ip_address))
        conn.commit()
    except Exception as e:
        logger.error(f"Error logging audit: {e}")
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

def check_account_locked(username):
    """Check if account is locked due to failed login attempts"""
    conn = get_rasa_connection()
    if not conn:
        return False
    
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT locked_until, failed_login_attempts 
            FROM users 
            WHERE username = %s
        """, (username,))
        
        result = cur.fetchone()
        if not result:
            return False
        
        locked_until, failed_attempts = result
        
        if locked_until and locked_until > datetime.now():
            return True
        
        # Reset lock if time has passed
        if locked_until and locked_until <= datetime.now():
            cur.execute("""
                UPDATE users 
                SET locked_until = NULL, failed_login_attempts = 0 
                WHERE username = %s
            """, (username,))
            conn.commit()
        
        return False
        
    except Exception as e:
        logger.error(f"Error checking account lock: {e}")
        return False
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

def increment_failed_login(username):
    """Increment failed login attempts and lock account if necessary"""
    conn = get_rasa_connection()
    if not conn:
        return
    
    try:
        cur = conn.cursor()
        cur.execute("""
            UPDATE users 
            SET failed_login_attempts = failed_login_attempts + 1
            WHERE username = %s
            RETURNING failed_login_attempts
        """, (username,))
        
        result = cur.fetchone()
        if result:
            attempts = result[0]
            
            # Lock account after 5 failed attempts for 30 minutes
            if attempts >= 5:
                lock_until = datetime.now() + timedelta(minutes=30)
                cur.execute("""
                    UPDATE users 
                    SET locked_until = %s 
                    WHERE username = %s
                """, (lock_until, username))
        
        conn.commit()
    except Exception as e:
        logger.error(f"Error incrementing failed login: {e}")
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

def reset_failed_login(username):
    """Reset failed login attempts on successful login"""
    conn = get_rasa_connection()
    if not conn:
        return
    
    try:
        cur = conn.cursor()
        cur.execute("""
            UPDATE users 
            SET failed_login_attempts = 0, 
                locked_until = NULL,
                last_login = CURRENT_TIMESTAMP
            WHERE username = %s
        """, (username,))
        conn.commit()
    except Exception as e:
        logger.error(f"Error resetting failed login: {e}")
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)
# ============================================
#  AUTHENTICATION ROUTES
# ============================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page and authentication"""
    # Redirect if already logged in
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            data = request.get_json() if request.is_json else request.form
            username = data.get('username', '').strip()
            password = data.get('password', '')
            
            if not username or not password:
                return jsonify({
                    'success': False,
                    'error': 'Username and password are required'
                }), 400
            
            # Check if account is locked
            if check_account_locked(username):
                log_audit(None, 'LOGIN_ATTEMPT_LOCKED', f'Username: {username}', request.remote_addr)
                return jsonify({
                    'success': False,
                    'error': 'Account is locked due to too many failed attempts. Please try again in 30 minutes.'
                }), 403
            
            conn = get_rasa_connection()
            if not conn:
                return jsonify({
                    'success': False,
                    'error': 'Database connection failed'
                }), 500
            
            try:
                cur = conn.cursor()
                cur.execute("""
                    SELECT id, username, password_hash, email, full_name, role, is_active
                    FROM users
                    WHERE username = %s
                """, (username,))
                
                user = cur.fetchone()
                
                if not user:
                    increment_failed_login(username)
                    log_audit(None, 'LOGIN_FAILED', f'Username: {username} (User not found)', request.remote_addr)
                    return jsonify({
                        'success': False,
                        'error': 'Invalid username or password'
                    }), 401
                
                user_id, db_username, password_hash, email, full_name, role, is_active = user
                
                # Check if account is active
                if not is_active:
                    log_audit(user_id, 'LOGIN_FAILED', 'Account disabled', request.remote_addr)
                    return jsonify({
                        'success': False,
                        'error': 'Account is disabled. Please contact administrator.'
                    }), 403
                
                # Verify password
                if not check_password_hash(password_hash, password):
                    increment_failed_login(username)
                    log_audit(user_id, 'LOGIN_FAILED', 'Invalid password', request.remote_addr)
                    return jsonify({
                        'success': False,
                        'error': 'Invalid username or password'
                    }), 401
                
                # Successful login
                reset_failed_login(username)
                
                # Set session
                session.permanent = True
                session['user_id'] = user_id
                session['username'] = db_username
                session['full_name'] = full_name or db_username
                session['email'] = email
                session['role'] = role
                
                # Log successful login
                log_audit(user_id, 'LOGIN_SUCCESS', f'Logged in as {role}', request.remote_addr)
                
                logger.info(f"✅ User logged in: {username} (ID: {user_id}, Role: {role})")
                
                return jsonify({
                    'success': True,
                    'message': 'Login successful',
                    'redirect': url_for('dashboard'),
                    'user': {
                        'username': db_username,
                        'full_name': full_name,
                        'role': role
                    }
                })
                
            finally:
                if 'cur' in locals():
                    cur.close()
                release_db_connection(conn, db_pool)
                
        except Exception as e:
            logger.error(f"Login error: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({
                'success': False,
                'error': 'An error occurred during login'
            }), 500
    
    # GET request - show login page
    return render_template('login.html')


# Cache statistics for 30 seconds
@lru_cache(maxsize=1)
def get_cached_stats(cache_key):
    """Get cached dashboard statistics"""
    complaints = get_unified_complaints()
    queue = get_agent_queue()
    
    # Calculate all stats
    return {
        'total_active': len(complaints),
        'in_queue': len(queue),
        'critical_count': sum(1 for c in complaints if c.get('priority') == 'CRITICAL'),
        # ... other stats
    }

# Invalidate cache every 30 seconds
def get_current_cache_key():
    return datetime.now().strftime('%Y%m%d%H%M%S')[:-1]  # Round to 10 seconds


@app.route('/api/dashboard_stats')
def api_dashboard_stats():
    """Get real-time dashboard statistics"""
    try:
        complaints = get_unified_complaints()
        queue = get_agent_queue()
        
        # Calculate stats
        today_start = datetime.combine(date.today(), time(0, 0))
        today_end = datetime.combine(date.today(), time(23, 59, 59))
        
        total_active = len(complaints)
        in_queue = len(queue)
        critical_count = sum(1 for c in complaints if c.get('priority', '').upper() == 'CRITICAL')
        resolved_today = sum(1 for c in complaints 
                            if c['status'] == 'RESOLVED' and 
                            c['full_data'].get('timestamp') and 
                            today_start <= c['full_data']['timestamp'] <= today_end)
        
        return jsonify({
            'success': True,
            'total_active': total_active,
            'in_queue': in_queue,
            'critical_count': critical_count,
            'resolved_today': resolved_today
        })
    except Exception as e:
        logger.error(f"Error getting dashboard stats: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

app.route('/api/bulk_update_status', methods=['POST'])
def bulk_update_status():
    """Update status for multiple complaints at once"""
    try:
        data = request.get_json()
        complaints = data.get('complaints', [])  # List of {table, record_id}
        new_status = data.get('status', '').strip().upper()
        
        if not new_status or not complaints:
            return jsonify({'success': False, 'error': 'Status and complaints required'})
        
        conn = get_rasa_connection()
        if not conn:
            return jsonify({'success': False, 'error': 'Database connection failed'})
        
        try:
            cur = conn.cursor()
            updated = 0
            
            for complaint in complaints:
                table = complaint.get('table')
                record_id = complaint.get('record_id')
                pk_column = 'report_id' if table == 'power_outage_reports' else 'id'
                
                cur.execute(
                    f"UPDATE {table} SET status = %s WHERE {pk_column} = %s",
                    (new_status, record_id)
                )
                updated += cur.rowcount
            
            conn.commit()
            
            return jsonify({
                'success': True,
                'message': f'Updated {updated} complaints',
                'updated_count': updated
            })
            
        finally:
            if 'cur' in locals():
                cur.close()
            release_db_connection(conn, db_pool)
            
    except Exception as e:
        logger.error(f"Bulk update error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export_complaints', methods=['GET'])
@login_required
def export_complaints():
    """Export filtered complaints to CSV"""
    import csv
    from io import StringIO
    from flask import make_response
    
    try:
        # Get current filters
        status_filter = request.args.get('status', 'All Status')
        priority_filter = request.args.get('priority', 'All Priorities')
        type_filter = request.args.get('type', 'All Types')
        search_term = request.args.get('search', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        time_from = request.args.get('time_from', '')
        time_to = request.args.get('time_to', '')
        
        # Get complaints with filters
        complaints = get_unified_complaints(
            status_filter, priority_filter, type_filter, search_term,
            False, date_from, date_to, time_from, time_to
        )
        
        # Create CSV
        si = StringIO()
        writer = csv.writer(si)
        
        # Write headers
        writer.writerow([
            'Time', 'Customer ID', 'Job Order ID', 'Type', 
            'Description', 'Priority', 'Status', 'Source'
        ])
        
        # Write data
        for c in complaints:
            writer.writerow([
                c.get('time', ''),
                c.get('customer_id', ''),
                c.get('job_order_id', ''),
                c.get('issue_type', ''),
                c.get('description', ''),
                c.get('priority', ''),
                c.get('status', ''),
                c.get('source', '')
            ])
        
        # Create response
        output = make_response(si.getvalue())
        output.headers["Content-Disposition"] = f"attachment; filename=complaints_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        output.headers["Content-type"] = "text/csv"
        
        return output
        
    except Exception as e:
        logger.error(f"Export error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/logout')
def logout():
    """Logout user"""
    user_id = session.get('user_id')
    username = session.get('username')
    
    if user_id:
        log_audit(user_id, 'LOGOUT', f'User logged out: {username}', request.remote_addr)
        logger.info(f"👋 User logged out: {username}")
    
    session.clear()
    flash('You have been logged out successfully', 'success')
    return redirect(url_for('login'))

def ensure_location_columns():
    """Ensure all complaint tables have required columns, including user_id"""
    conn = get_rasa_connection()
    if not conn:
        logger.error("Failed to connect to database in ensure_location_columns")
        return False
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS power_outage_reports (
                report_id SERIAL PRIMARY KEY,
                user_id VARCHAR(100),
                full_name VARCHAR(100) NOT NULL,
                address TEXT NOT NULL,
                contact_number VARCHAR(15) NOT NULL,
                job_order_id VARCHAR(20) UNIQUE,
                latitude DOUBLE PRECISION,
                longitude DOUBLE PRECISION,
                accuracy DOUBLE PRECISION,
                issue_type VARCHAR(50) NOT NULL,
                priority VARCHAR(20) NOT NULL,
                status VARCHAR(20) NOT NULL,
                source VARCHAR(50) NOT NULL,
                timestamp TIMESTAMP NOT NULL,
                details TEXT,
                hidden BOOLEAN DEFAULT FALSE
            )
        """)
        cur.execute("""
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name = 'power_outage_reports' AND column_name = 'user_id'
                ) THEN
                    ALTER TABLE power_outage_reports ADD COLUMN user_id VARCHAR(100);
                END IF;
            END $$;
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS meter_concern (
                id SERIAL PRIMARY KEY,
                user_id VARCHAR(100),
                job_order_id VARCHAR(20) UNIQUE,
                account_no VARCHAR(20) NOT NULL,
                name VARCHAR(100) NOT NULL,
                address TEXT NOT NULL,
                contact_number VARCHAR(15) NOT NULL,
                concern TEXT NOT NULL,
                priority VARCHAR(20) NOT NULL,
                status VARCHAR(20) NOT NULL,
                source VARCHAR(50) NOT NULL,
                timestamp TIMESTAMP NOT NULL,
                latitude DOUBLE PRECISION,
                longitude DOUBLE PRECISION,
                accuracy DOUBLE PRECISION,
                hidden BOOLEAN DEFAULT FALSE
            )
        """)
        cur.execute("""
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name = 'meter_concern' AND column_name = 'user_id'
                ) THEN
                    ALTER TABLE meter_concern ADD COLUMN user_id VARCHAR(100);
                END IF;
            END $$;
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS agent_queue (
                id SERIAL PRIMARY KEY,
                user_id VARCHAR(100) NOT NULL,
                full_name VARCHAR(100) NOT NULL,
                concern TEXT NOT NULL,
                contact_number VARCHAR(15) NOT NULL,
                priority VARCHAR(20) NOT NULL,
                timestamp TIMESTAMP NOT NULL,
                status VARCHAR(20) NOT NULL,
                latitude DOUBLE PRECISION,
                longitude DOUBLE PRECISION,
                accuracy DOUBLE PRECISION,
                hidden BOOLEAN DEFAULT FALSE,
                resumed BOOLEAN DEFAULT FALSE,
                job_order_id VARCHAR(50)
            )
        """)
        conn.commit()
        logger.info("Successfully ensured location columns for all tables")
        return True
    except Exception as e:
        logger.error(f"Error creating tables: {e}")
        conn.rollback()
        return False
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

def ensure_issue_type_column():
    """Ensure issue_type column exists in power_outage_reports"""
    conn = get_rasa_connection()
    if not conn:
        logger.error("Failed to connect to database in ensure_issue_type_column")
        return False
    try:
        cur = conn.cursor()
        cur.execute("""
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name = 'power_outage_reports' AND column_name = 'issue_type'
                ) THEN
                    ALTER TABLE power_outage_reports ADD COLUMN issue_type VARCHAR(50) DEFAULT 'Power Outage';
                END IF;
                UPDATE power_outage_reports SET issue_type = 'Power Outage' WHERE issue_type IS NULL;
            END $$;
        """)
        conn.commit()
        logger.info("Successfully ensured issue_type column")
        return True
    except Exception as e:
        logger.error(f"Error ensuring issue_type column: {e}")
        conn.rollback()
        return False
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

def enable_postgis_extension(conn):
    """Enable PostGIS extension in the database"""
    try:
        cur = conn.cursor()
        # Enable PostGIS extension
        cur.execute("CREATE EXTENSION IF NOT EXISTS postgis")
        conn.commit()
        logger.info("✓ PostGIS extension enabled")
        return True
    except Exception as e:
        logger.error(f"✗ Error enabling PostGIS: {e}")
        conn.rollback()
        return False
    finally:
        if 'cur' in locals():
            cur.close()

def migrate_to_geometry():
    """
    Migrate existing lat/lng columns to PostGIS geometry type
    SRID 4326 = WGS84 (standard GPS coordinates)
    """
    conn = get_rasa_connection()
    if not conn:
        logger.error("Database connection failed for geometry migration")
        return False
    
    try:
        cur = conn.cursor()
        
        # Enable PostGIS
        enable_postgis_extension(conn)
        
        tables = ['power_outage_reports', 'meter_concern', 'agent_queue']
        
        for table in tables:
            logger.info(f"Migrating {table} to geometry...")
            
            # Check if table exists
            cur.execute("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_schema = 'public' AND table_name = %s
                )
            """, (table,))
            
            if not cur.fetchone()[0]:
                logger.warning(f"Table {table} does not exist, skipping")
                continue
            
            # Add geometry column (POINT type, SRID 4326 for GPS coordinates)
            cur.execute(f"""
                DO $$
                BEGIN
                    -- Add geometry column if it doesn't exist
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name = '{table}' AND column_name = 'geom'
                    ) THEN
                        ALTER TABLE {table} 
                        ADD COLUMN geom GEOMETRY(POINT, 4326);
                        
                        RAISE NOTICE 'Added geometry column to {table}';
                    END IF;
                END $$;
            """)
            
            # Populate geometry from existing lat/lng
            cur.execute(f"""
                UPDATE {table}
                SET geom = ST_SetSRID(ST_MakePoint(longitude, latitude), 4326)
                WHERE latitude IS NOT NULL 
                  AND longitude IS NOT NULL
                  AND geom IS NULL;
            """)
            
            rows_updated = cur.rowcount
            logger.info(f"  ✓ Updated {rows_updated} rows with geometry data")
            
            # Create spatial index for fast queries
            cur.execute(f"""
                CREATE INDEX IF NOT EXISTS {table}_geom_idx 
                ON {table} USING GIST (geom);
            """)
            logger.info(f"  ✓ Created spatial index on {table}")
            
            # Add trigger to auto-update geometry when lat/lng changes
            cur.execute(f"""
                CREATE OR REPLACE FUNCTION update_{table}_geometry()
                RETURNS TRIGGER AS $$
                BEGIN
                    IF NEW.latitude IS NOT NULL AND NEW.longitude IS NOT NULL THEN
                        NEW.geom = ST_SetSRID(ST_MakePoint(NEW.longitude, NEW.latitude), 4326);
                    END IF;
                    RETURN NEW;
                END;
                $$ LANGUAGE plpgsql;
                
                DROP TRIGGER IF EXISTS trigger_update_{table}_geometry ON {table};
                
                CREATE TRIGGER trigger_update_{table}_geometry
                BEFORE INSERT OR UPDATE OF latitude, longitude
                ON {table}
                FOR EACH ROW
                EXECUTE FUNCTION update_{table}_geometry();
            """)
            logger.info(f"  ✓ Created auto-update trigger for {table}")
        
        conn.commit()
        logger.info("✅ Geometry migration completed successfully!")
        
        # Verify migration
        for table in tables:
            cur.execute(f"""
                SELECT COUNT(*) as total,
                       COUNT(geom) as with_geometry
                FROM {table}
            """)
            result = cur.fetchone()
            logger.info(f"  {table}: {result[1]}/{result[0]} records have geometry")
        
        return True
        
    except Exception as e:
        logger.error(f"❌ Error during geometry migration: {e}")
        import traceback
        traceback.print_exc()
        conn.rollback()
        return False
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

def migrate_incident_details_columns():
    """
    Add incident detail columns to power_outage_reports table
    Run this function once to add the new columns
    """
    conn = get_rasa_connection()
    if not conn:
        logger.error("Failed to connect to database for incident details migration")
        return False
    
    try:
        cur = conn.cursor()
        
        logger.info("🔧 Adding incident details columns to power_outage_reports...")
        
        # Add all the incident detail columns
        cur.execute("""
            DO $$
            BEGIN
                -- Add incident_type_detail column (power_outage, fallen_wire, etc.)
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name = 'power_outage_reports' AND column_name = 'incident_type_detail'
                ) THEN
                    ALTER TABLE power_outage_reports ADD COLUMN incident_type_detail VARCHAR(100);
                    RAISE NOTICE 'Added incident_type_detail column';
                END IF;
                
                -- Add affected_area column (single_house, multiple_houses, etc.)
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name = 'power_outage_reports' AND column_name = 'affected_area'
                ) THEN
                    ALTER TABLE power_outage_reports ADD COLUMN affected_area VARCHAR(100);
                    RAISE NOTICE 'Added affected_area column';
                END IF;
                
                -- Add incident_time column
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name = 'power_outage_reports' AND column_name = 'incident_time'
                ) THEN
                    ALTER TABLE power_outage_reports ADD COLUMN incident_time TIME;
                    RAISE NOTICE 'Added incident_time column';
                END IF;
                
                -- Add duration column
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name = 'power_outage_reports' AND column_name = 'duration'
                ) THEN
                    ALTER TABLE power_outage_reports ADD COLUMN duration VARCHAR(50);
                    RAISE NOTICE 'Added duration column';
                END IF;
                
                -- Add landmark column
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name = 'power_outage_reports' AND column_name = 'landmark'
                ) THEN
                    ALTER TABLE power_outage_reports ADD COLUMN landmark TEXT;
                    RAISE NOTICE 'Added landmark column';
                END IF;
                
                -- Add email column (was in form but may not be in DB)
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name = 'power_outage_reports' AND column_name = 'email'
                ) THEN
                    ALTER TABLE power_outage_reports ADD COLUMN email VARCHAR(255);
                    RAISE NOTICE 'Added email column';
                END IF;
                
                -- Add account_number column (was in form but may not be in DB)
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name = 'power_outage_reports' AND column_name = 'account_number'
                ) THEN
                    ALTER TABLE power_outage_reports ADD COLUMN account_number VARCHAR(50);
                    RAISE NOTICE 'Added account_number column';
                END IF;
            END $$;
        """)
        
        conn.commit()
        logger.info("✅ Incident details columns added successfully!")
        
        # Verify the columns were added
        cur.execute("""
            SELECT column_name, data_type 
            FROM information_schema.columns 
            WHERE table_name = 'power_outage_reports' 
            AND column_name IN (
                'incident_type_detail', 'affected_area', 'incident_time', 
                'duration', 'landmark', 'email', 'account_number'
            )
            ORDER BY column_name
        """)
        
        columns = cur.fetchall()
        logger.info("✅ Verified columns:")
        for col in columns:
            logger.info(f"   - {col[0]} ({col[1]})")
        
        return True
        
    except Exception as e:
        logger.error(f"❌ Error adding incident details columns: {e}")
        import traceback
        traceback.print_exc()
        conn.rollback()
        return False
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

def get_complaints_near_location(lat, lng, radius_meters=1000):
    """
    Find complaints within radius of a location using PostGIS
    
    Args:
        lat: Latitude
        lng: Longitude
        radius_meters: Search radius in meters (default 1km)
    
    Returns:
        List of complaints within radius
    """
    conn = get_rasa_connection()
    if not conn:
        return []
    
    try:
        cur = conn.cursor()
        
        # Search all tables for nearby complaints
        complaints = []
        
        # Power outage reports
        cur.execute("""
            SELECT 
                report_id,
                full_name,
                address,
                ST_X(geom) as lng,
                ST_Y(geom) as lat,
                ST_Distance(
                    geom::geography,
                    ST_SetSRID(ST_MakePoint(%s, %s), 4326)::geography
                ) as distance_meters,
                status,
                priority,
                timestamp
            FROM power_outage_reports
            WHERE geom IS NOT NULL
              AND ST_DWithin(
                  geom::geography,
                  ST_SetSRID(ST_MakePoint(%s, %s), 4326)::geography,
                  %s
              )
              AND (hidden = FALSE OR hidden IS NULL)
            ORDER BY distance_meters
        """, (lng, lat, lng, lat, radius_meters))
        
        for row in cur.fetchall():
            complaints.append({
                'id': f"PO-{row[0]:06d}",
                'customer': row[1],
                'address': row[2],
                'lng': row[3],
                'lat': row[4],
                'distance_meters': round(row[5], 2),
                'distance_km': round(row[5] / 1000, 2),
                'status': row[6],
                'priority': row[7],
                'timestamp': row[8],
                'type': 'Power Outage'
            })
        
        logger.info(f"Found {len(complaints)} complaints within {radius_meters}m of ({lat}, {lng})")
        return complaints
        
    except Exception as e:
        logger.error(f"Error finding nearby complaints: {e}")
        return []
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)



def get_complaint_clusters(zoom_level=12):
    """
    Get clustered complaint locations for map display
    Uses PostGIS ST_ClusterKMeans for intelligent clustering
    """
    conn = get_rasa_connection()
    if not conn:
        return []
    
    try:
        cur = conn.cursor()
        
        # Determine cluster size based on zoom level
        num_clusters = min(50, max(5, 100 // (zoom_level + 1)))
        
        cur.execute("""
            WITH clustered AS (
                SELECT 
                    report_id,
                    full_name,
                    status,
                    priority,
                    ST_ClusterKMeans(geom, %s) OVER () as cluster_id,
                    geom
                FROM power_outage_reports
                WHERE geom IS NOT NULL
                  AND (hidden = FALSE OR hidden IS NULL)
            )
            SELECT 
                cluster_id,
                COUNT(*) as complaint_count,
                ST_X(ST_Centroid(ST_Collect(geom))) as center_lng,
                ST_Y(ST_Centroid(ST_Collect(geom))) as center_lat,
                array_agg(report_id) as complaint_ids,
                array_agg(status) as statuses,
                array_agg(priority) as priorities
            FROM clustered
            GROUP BY cluster_id
            ORDER BY complaint_count DESC
        """, (num_clusters,))
        
        clusters = []
        for row in cur.fetchall():
            clusters.append({
                'cluster_id': row[0],
                'count': row[1],
                'center_lng': row[2],
                'center_lat': row[3],
                'complaint_ids': row[4],
                'statuses': row[5],
                'priorities': row[6]
            })
        
        return clusters
        
    except Exception as e:
        logger.error(f"Error getting complaint clusters: {e}")
        return []
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

@app.route('/api/complaints')
def get_all_complaints():
    """Fetch all complaints from multiple tables"""
    try:
        conn = psycopg2.connect(
            host="localhost",
            database="rasa_db",
            user="postgres",
            password="5227728"
        )
        cursor = conn.cursor()
        
        complaints = []
        
        # Fetch from power_outage_reports
        cursor.execute("""
            SELECT 
                job_order_id,
                full_name as customer_id,
                address,
                contact_number,
                issue_type,
                priority,
                status,
                source,
                TO_CHAR(timestamp, 'YYYY-MM-DD HH24:MI:SS') as time,
                description,
                'power_outage_reports' as table,
                id as record_id
            FROM power_outage_reports
            WHERE status != 'RESOLVED'
            ORDER BY timestamp DESC
        """)
        
        for row in cursor.fetchall():
            complaints.append({
                'job_order_id': row[0],
                'customer_id': row[1],
                'address': row[2],
                'contact_number': row[3],
                'issue_type': row[4],
                'priority': row[5],
                'status': row[6],
                'source': row[7],
                'time': row[8],
                'description': row[9],
                'table': row[10],
                'record_id': row[11]
            })
        
        # Fetch from meter_concern
        cursor.execute("""
            SELECT 
                job_order_id,
                name as customer_id,
                address,
                contact_number,
                'SERVICE' as issue_type,
                priority,
                status,
                source,
                TO_CHAR(timestamp, 'YYYY-MM-DD HH24:MI:SS') as time,
                concern as description,
                'meter_concern' as table,
                id as record_id
            FROM meter_concern
            WHERE status != 'RESOLVED'
            ORDER BY timestamp DESC
        """)
        
        for row in cursor.fetchall():
            complaints.append({
                'job_order_id': row[0],
                'customer_id': row[1],
                'address': row[2],
                'contact_number': row[3],
                'issue_type': row[4],
                'priority': row[5],
                'status': row[6],
                'source': row[7],
                'time': row[8],
                'description': row[9],
                'table': row[10],
                'record_id': row[11]
            })
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'complaints': complaints,
            'count': len(complaints)
        })
        
    except Exception as e:
        print(f"Error fetching complaints: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def ensure_timestamp_column():
    """Ensure timestamp column in all tables is TIMESTAMP type"""
    conn = get_rasa_connection()
    if not conn:
        logger.error("Failed to connect to database in ensure_timestamp_column")
        return False
    try:
        cur = conn.cursor()
        tables = ['power_outage_reports', 'meter_concern', 'agent_queue']
        for table in tables:
            cur.execute(f"""
                DO $$
                BEGIN
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name = '{table}' AND column_name = 'timestamp'
                    ) THEN
                        ALTER TABLE {table} ADD COLUMN timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP;
                    END IF;
                    IF EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name = '{table}' AND column_name = 'timestamp'
                        AND data_type != 'timestamp without time zone'
                    ) THEN
                        ALTER TABLE {table} ADD COLUMN timestamp_temp TIMESTAMP;
                        UPDATE {table} 
                        SET timestamp_temp = CASE 
                            WHEN timestamp IS NOT NULL AND timestamp ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2} [0-9]{2}:[0-9]{2}:[0-9]{2}$' 
                            THEN to_timestamp(timestamp, 'YYYY-MM-DD HH24:MI:SS')
                            ELSE CURRENT_TIMESTAMP
                        END;
                        ALTER TABLE {table} DROP COLUMN timestamp;
                        ALTER TABLE {table} RENAME COLUMN timestamp_temp TO timestamp;
                    END IF;
                    UPDATE {table} SET timestamp = CURRENT_TIMESTAMP 
                    WHERE timestamp IS NULL OR timestamp::text !~ '^[0-9]{4}-[0-9]{2}-[0-9]{2} [0-9]{2}:[0-9]{2}:[0-9]{2}$';
                END $$;
            """)
        conn.commit()
        logger.info("Successfully ensured timestamp columns")
        return True
    except Exception as e:
        logger.error(f"Error ensuring timestamp column: {e}")
        conn.rollback()
        return False
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

@app.route('/api/complaints')
def get_complaints():
    try:
        conn = psycopg2.connect(
            host="localhost",
            database="rasa_db",
            user="postgres",
            password="5227728"
        )
        cursor = conn.cursor()
        
        # ✅ Make sure you're querying the right table
        cursor.execute("""
            SELECT 
                job_order_id,
                full_name as customer,
                address,
                contact_number as contact,
                issue_type,
                priority,
                status,
                source,
                timestamp as time,
                'power_outage_reports' as table,
                id as record_id
            FROM power_outage_reports
            WHERE status != 'RESOLVED'
            ORDER BY timestamp DESC
        """)
        
        complaints = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'complaints': complaints})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# New API endpoint for nearby search
@app.route('/api/complaints_nearby', methods=['POST'])
def api_complaints_nearby():
    """Find complaints near a location"""
    try:
        data = request.get_json()
        lat = float(data.get('lat'))
        lng = float(data.get('lng'))
        radius = int(data.get('radius', 1000))  # Default 1km
        
        complaints = get_complaints_near_location(lat, lng, radius)
        
        return jsonify({
            'success': True,
            'complaints': complaints,
            'count': len(complaints),
            'search_center': {'lat': lat, 'lng': lng},
            'radius_meters': radius
        })
        
    except Exception as e:
        logger.error(f"Error in nearby complaints API: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

def get_priority_from_concern(concern: str) -> str:
    """
    Classify the priority based on concern description
    - Critical: Life-threatening, hazardous, or emergency situations
    - High: Major outages affecting service
    """
    concern = (concern or "").lower()
    
    # Critical (🚨 Life-threatening or hazardous - IMMEDIATE RESPONSE)
    if any(word in concern for word in [
        "fire", "explosion", "burning", "smoke", "accident", 
        "fallen wire", "electric shock", "live wire", "transformer burst",
        "emergency", "danger", "hazard", "sparking", "exposed wire",
        "electrocuted", "injured", "death", "pole down", "wire down",
        "short circuit", "arcing", "flames"
    ]):
        return "CRITICAL"
    
    # Everything else defaults to HIGH for power outage reports
    # This includes: power outages, blackouts, no electricity, flickering, voltage issues
    return "HIGH"

def create_indexes():
    """Create indexes on frequently queried columns"""
    conn = get_rasa_connection()
    if not conn:
        logger.error("Failed to connect to database for index creation")
        return False
    try:
        cur = conn.cursor()
        cur.execute("CREATE INDEX IF NOT EXISTS idx_power_outage_timestamp ON power_outage_reports(timestamp)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_power_outage_status ON power_outage_reports(status)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_power_outage_hidden ON power_outage_reports(hidden)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_meter_concern_timestamp ON meter_concern(timestamp)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_agent_queue_timestamp ON agent_queue(timestamp)")
        conn.commit()
        logger.info("Database indexes created successfully")
        return True
    except Exception as e:
        logger.error(f"Error creating indexes: {e}")
        conn.rollback()
        return False
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

def ensure_converted_table_joblist():
    """Ensure the converted table exists in joblist database"""
    conn = get_joblist_connection()
    if not conn:
        logger.error("Failed to connect to joblist database")
        return False
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS public.converted (
                unique_id TEXT,
                creator TEXT,
                created TEXT,
                follower TEXT,
                followed TEXT,
                name TEXT,
                spinners TEXT,
                town0 TEXT,
                brgy0 TEXT,
                town TEXT,
                brgy TEXT,
                town2 TEXT,
                brgy2 TEXT,
                assignedto TEXT,
                status TEXT,
                subs TEXT,
                feeder TEXT,
                section TEXT,
                cause TEXT,
                equip TEXT,
                type TEXT,
                notes TEXT,
                landmark TEXT,
                phone TEXT,
                location TEXT,
                latitude TEXT,
                longitude TEXT,
                actiontaken TEXT
            )
        """)
        conn.commit()
        logger.info("Successfully ensured converted table in joblist DB")
        return True
    except Exception as e:
        logger.error(f"Error ensuring converted table in joblist DB: {e}")
        conn.rollback()
        return False
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, joblist_db_pool)



@app.route('/api/submit_power_outage', methods=['POST'])
def submit_power_outage():
    """Submit power outage report with incident details and PostGIS geometry"""
    try:
        data = request.get_json()
        
        # Extract ALL form data including incident details
        full_name = data.get('full_name', '').strip()
        contact_number = data.get('contact_number', '').strip()
        email = data.get('email', '').strip()
        account_number = data.get('account_number', '').strip()
        address = data.get('address', '').strip()
        latitude = data.get('latitude')
        longitude = data.get('longitude')
        accuracy = data.get('accuracy')
        
        # Extract incident details
        incident_type = data.get('incident_type', 'power_outage')
        affected_area = data.get('affected_area', 'unknown')
        incident_time = data.get('incident_time', '')
        duration = data.get('duration', 'just_now')
        details = data.get('details', '').strip()
        landmark = data.get('landmark', '').strip()
        source = data.get('source', 'Web Form')
        
        # Validate required fields
        if not all([full_name, contact_number, address, details, latitude, longitude]):
            return jsonify({
                'success': False,
                'error': 'Missing required fields'
            }), 400
        
        # Determine priority
        critical_types = ['fallen_wire', 'fire_hazard', 'transformer_issue']
        if incident_type in critical_types:
            priority = 'CRITICAL'
        else:
            priority = get_priority_from_concern(details)
        
        logger.info(f"📝 Submitting report with incident details: {full_name} at ({latitude}, {longitude})")
        logger.info(f"   Incident Type: {incident_type}, Affected Area: {affected_area}")
        
        conn = get_rasa_connection()
        if not conn:
            return jsonify({
                'success': False,
                'error': 'Database connection failed'
            }), 500
        
        cur = conn.cursor()
        
        # Insert with ALL incident detail fields
        insert_query = """
            INSERT INTO power_outage_reports 
            (full_name, contact_number, email, account_number, address, 
             latitude, longitude, accuracy, details, landmark,
             incident_type_detail, affected_area, incident_time, duration,
             priority, status, source, issue_type, timestamp)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
            RETURNING report_id, ST_AsText(geom) as geom_wkt, ST_X(geom) as lng, ST_Y(geom) as lat
        """
        
        cur.execute(insert_query, (
            full_name,
            contact_number,
            email,
            account_number,
            address,
            latitude,
            longitude,
            accuracy,
            details,
            landmark,
            incident_type,
            affected_area,
            incident_time,
            duration,
            priority,
            'NEW',
            source,
            'Power Outage'
        ))
        
        result = cur.fetchone()
        report_id = result[0]
        geom_wkt = result[1]
        geom_lng = result[2]
        geom_lat = result[3]
        
        logger.info(f"✅ Report {report_id} created with incident details and geometry: {geom_wkt}")
        
        conn.commit()
        
        # ============================================
        # 🔴 ADD THIS: BROADCAST NEW COMPLAINT
        # ============================================
        complaint_data = {
            'record_id': report_id,
            'customer_name': full_name,
            'issue_type': 'Power Outage',
            'priority': priority,
            'incident_type': incident_type,
            'affected_area': affected_area,
            'address': address,
            'timestamp': datetime.now().isoformat()
        }
        
        # Broadcast new complaint to all connected clients
        broadcast_new_complaint(complaint_data)
        
        # If critical, send alert
        if priority == 'CRITICAL':
            broadcast_critical_alert(complaint_data)
        # ============================================
        
        # Create job order for CRITICAL and HIGH priority
        job_order_id = None
        if priority in ['CRITICAL', 'HIGH']:
            try:
                brgy0, town0 = extract_location_parts(address)
                converted_unique_id = generate_unique_id()
                
                converted_data = {
                    'unique_id': converted_unique_id,
                    'creator': contact_number or full_name,
                    'created': datetime.now().strftime('%m/%d/%y %I:%M:%S %p'),
                    'follower': contact_number or full_name,
                    'followed': datetime.now().strftime('%m/%d/%y %I:%M:%S %p'),
                    'name': full_name,
                    'spinners': contact_number,
                    'town0': town0,
                    'brgy0': brgy0,
                    'town': 'Select Town',
                    'brgy': 'Select Brgy',
                    'town2': 'Select Town',
                    'brgy2': 'Select Brgy',
                    'assignedto': town0 or '',
                    'status': 'Select Status',
                    'subs': 'Substation',
                    'feeder': 'Feeder',
                    'section': 'Category',
                    'cause': f"{incident_type.replace('_', ' ').title()}: {details}",
                    'equip': 'Equipment',
                    'type': 'high' if priority == 'CRITICAL' else 'high',
                    'notes': f"Priority: {priority} | Type: {incident_type} | Area: {affected_area} | {details}",
                    'landmark': landmark or '',
                    'phone': contact_number,
                    'location': address,
                    'latitude': str(latitude),
                    'longitude': str(longitude),
                    'actiontaken': 'Pending'
                }
                
                joblist_conn = get_joblist_connection()
                if joblist_conn:
                    joblist_cur = joblist_conn.cursor()
                    columns = list(converted_data.keys())
                    values = list(converted_data.values())
                    placeholders = ', '.join(['%s'] * len(values))
                    
                    insert_query = f"INSERT INTO converted ({', '.join(columns)}) VALUES ({placeholders})"
                    joblist_cur.execute(insert_query, values)
                    joblist_conn.commit()
                    joblist_cur.close()
                    release_db_connection(joblist_conn, joblist_db_pool)
                    
                    job_order_id = converted_unique_id
                    
                    # Update report with job order ID
                    cur.execute(
                        "UPDATE power_outage_reports SET job_order_id = %s WHERE report_id = %s",
                        (job_order_id, report_id)
                    )
                    conn.commit()
                    
                    logger.info(f"✅ Job order {job_order_id} created for report {report_id}")
                    
            except Exception as e:
                logger.error(f"Error creating automatic job order: {e}")
                traceback.print_exc()
        
        cur.close()
        release_db_connection(conn, db_pool)
        
        return jsonify({
            'success': True,
            'report_id': report_id,
            'job_order_id': job_order_id or 'Will be assigned by operator',
            'priority': priority,
            'latitude': geom_lat,
            'longitude': geom_lng,
            'geometry': geom_wkt,
            'incident_details': {
                'type': incident_type,
                'affected_area': affected_area,
                'incident_time': incident_time,
                'duration': duration
            },
            'message': f'Power outage report submitted successfully with incident details'
        })
        
    except Exception as e:
        logger.error(f"Error submitting power outage: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
# Run migration on startup
def initialize_postgis():
    """Initialize PostGIS support on application startup"""
    logger.info("=" * 60)
    logger.info("🗺️  Initializing PostGIS Geometry Support")
    logger.info("=" * 60)
    
    try:
        if migrate_to_geometry():
            logger.info("✅ PostGIS initialization complete")
            return True
        else:
            logger.warning("⚠️  PostGIS initialization failed - using lat/lng fallback")
            return False
    except Exception as e:
        logger.error(f"❌ PostGIS initialization error: {e}")
        return False

def parse_datetime_filters(date_from=None, date_to=None, time_from=None, time_to=None):
    """Parse and combine date and time filters into datetime objects"""
    start_datetime = None
    end_datetime = None
    try:
        if date_from:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            start_time = time(0, 0)
            if time_from:
                start_time = datetime.strptime(time_from, '%H:%M').time()
            start_datetime = datetime.combine(start_date, start_time)
        if date_to:
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            end_time = time(23, 59, 59)
            if time_to:
                end_time = datetime.strptime(time_to, '%H:%M').time()
            end_datetime = datetime.combine(end_date, end_time)
    except ValueError as e:
        logger.error(f"Error parsing date/time filters: {e}")
    return start_datetime, end_datetime

def get_agent_queue():
    """Get agent queue data from rasa_db"""
    conn = get_rasa_connection()
    if not conn:
        logger.error("Failed to connect to database for agent queue")
        return []
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT id, user_id, full_name, concern, contact_number, priority, timestamp, 
                   COALESCE(status, 'NEW') AS status, latitude, longitude, accuracy
            FROM agent_queue
            WHERE (hidden = FALSE OR hidden IS NULL) AND (resumed = FALSE OR resumed IS NULL)
            ORDER BY timestamp ASC
        """)
        rows = cur.fetchall()
        result = [{
            'id': f"AQ-{row[0]:06d}",
            'customer': row[2] or 'Unknown',
            'customerId': row[1] or 'N/A',
            'address': None,
            'lat': row[8],
            'lng': row[9],
            'issueType': 'Service',
            'description': row[3] or 'Service request',
            'priority': (row[5] or 'LOW').upper(),
            'status': (row[7] or 'NEW').upper(),
            'time': row[6].strftime('%I:%M %p') if row[6] else 'N/A',
            'contact': row[4] or 'N/A',
            'source': 'Agent Queue',
            'accuracy': row[10],
            'details': row[3] or 'No details provided'
        } for row in rows]
        logger.info(f"Retrieved {len(result)} agent queue records")
        return result
    except Exception as e:
        logger.error(f"DB error in get_agent_queue: {e}")
        return []
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

def ensure_columns(conn, table):
    """Ensure status, hidden, and resumed columns exist"""
    try:
        cur = conn.cursor()
        cur.execute(f"""
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name = '{table}' AND column_name = 'status'
                ) THEN
                    ALTER TABLE {table} ADD COLUMN status VARCHAR(50) DEFAULT 'NEW';
                END IF;
            END $$;
        """)
        cur.execute(f"""
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name = '{table}' AND column_name = 'hidden'
                ) THEN
                    ALTER TABLE {table} ADD COLUMN hidden BOOLEAN DEFAULT FALSE;
                END IF;
            END $$;
        """)
        if table == 'agent_queue':
            cur.execute(f"""
                DO $$
                BEGIN
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name = 'agent_queue' AND column_name = 'resumed'
                    ) THEN
                        ALTER TABLE agent_queue ADD COLUMN resumed BOOLEAN DEFAULT FALSE;
                    END IF;
                END $$;
            """)
            cur.execute(f"""
                DO $$
                BEGIN
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name = 'agent_queue' AND column_name = 'job_order_id'
                    ) THEN
                        ALTER TABLE agent_queue ADD COLUMN job_order_id VARCHAR(50);
                    END IF;
                END $$;
            """)
        conn.commit()
        logger.info(f"Ensured columns for table {table}")
        return True
    except Exception as e:
        logger.error(f"Error ensuring columns for {table}: {e}")
        conn.rollback()
        return False
    finally:
        if 'cur' in locals():
            cur.close()

# Improved get_unified_complaints function with better error handling
def get_unified_complaints(status_filter=None, priority_filter=None, type_filter=None, search_term=None,
                           show_hidden=False, date_from=None, date_to=None, time_from=None, time_to=None):
    """Get unified complaints from all tables with filters"""
    complaints = []
    
    # Normalize filters
    status_filter_norm = status_filter.strip().upper() if status_filter and status_filter != 'All Status' else None
    priority_filter_norm = priority_filter.strip().upper() if priority_filter and priority_filter != 'All Priorities' else None
    type_filter_norm = type_filter.strip().upper() if type_filter and type_filter != 'All Types' else None
    
    logger.info(f"🔍 Fetching complaints with filters - Status: {status_filter_norm}, Priority: {priority_filter_norm}, Type: {type_filter_norm}, Search: {search_term}")
    
    start_datetime, end_datetime = parse_datetime_filters(date_from, date_to, time_from, time_to)
    
    conn = get_rasa_connection()
    if not conn:
        logger.error("❌ Database connection failed for unified complaints")
        return []

    try:
        cur = conn.cursor()
        
        # Get list of existing tables
        cur.execute("""
            SELECT table_name FROM information_schema.tables 
            WHERE table_schema = 'public' AND table_name IN ('power_outage_reports', 'meter_concern', 'agent_queue')
        """)
        existing_tables = [row[0] for row in cur.fetchall()]
        logger.info(f"📋 Found tables: {existing_tables}")

        # Process each table
        for table in ['agent_queue', 'meter_concern', 'power_outage_reports']:
            if table not in existing_tables:
                logger.warning(f"⚠️ Table {table} does not exist, skipping")
                continue
            
            # Ensure columns exist
            ensure_columns(conn, table)
            
            try:
                if table == 'agent_queue':
                    query = """
                        SELECT id, user_id, full_name, concern, contact_number, 
                               COALESCE(priority, 'LOW') as priority, 
                               timestamp, 
                               COALESCE(status, 'NEW') as status, 
                               latitude, longitude, accuracy, hidden, resumed
                        FROM agent_queue
                        WHERE 1=1
                    """
                    params = []
                    
                    if not show_hidden:
                        query += " AND (hidden = FALSE OR hidden IS NULL)"
                    
                    query += " AND (resumed = FALSE OR resumed IS NULL)"
                    
                    if search_term:
                        query += " AND (user_id ILIKE %s OR full_name ILIKE %s OR concern ILIKE %s)"
                        search_pattern = f"%{search_term}%"
                        params.extend([search_pattern, search_pattern, search_pattern])
                    
                    if status_filter_norm:
                        query += " AND UPPER(COALESCE(status, 'NEW')) = %s"
                        params.append(status_filter_norm)
                    
                    if priority_filter_norm:
                        query += " AND UPPER(COALESCE(priority, 'LOW')) = %s"
                        params.append(priority_filter_norm)
                    
                    if type_filter_norm and type_filter_norm != 'SERVICE':
                        continue  # Skip this table if type filter doesn't match
                    
                    if start_datetime:
                        query += " AND timestamp >= %s"
                        params.append(start_datetime)
                    
                    if end_datetime:
                        query += " AND timestamp <= %s"
                        params.append(end_datetime)
                    
                    query += " ORDER BY timestamp DESC"
                    
                    logger.debug(f"Agent Queue Query: {query}")
                    logger.debug(f"Params: {params}")
                    
                    cur.execute(query, params)
                    rows = cur.fetchall()
                    logger.info(f"✅ Found {len(rows)} agent queue records")
                    
                    for row in rows:
                        complaints.append({
                            'time': row[6].strftime('%I:%M %p') if row[6] else 'N/A',
                            'customer_id': row[1] or 'N/A',
                            'job_order_id': None,
                            'issue_type': 'Service',
                            'description': (row[3][:50] + '...') if row[3] and len(row[3]) > 50 else (row[3] or 'Service request'),
                            'priority': (row[5] or 'LOW').upper(),
                            'status': (row[7] or 'NEW').upper(),
                            'source': 'Agent Request',
                            'table': 'agent_queue',
                            'record_id': row[0],
                            'customer_name': row[2] or 'Unknown',
                            'customer_phone': row[4] or 'N/A',
                            'full_data': {
                                'id': row[0],
                                'user_id': row[1],
                                'full_name': row[2],
                                'concern': row[3],
                                'contact_number': row[4],
                                'priority': row[5],
                                'timestamp': row[6],
                                'status': row[7],
                                'latitude': row[8],
                                'longitude': row[9],
                                'accuracy': row[10]
                            }
                        })
                
                elif table == 'meter_concern':
                    query = """
                        SELECT id, user_id, account_no, name, address, contact_number, concern, 
                               timestamp, job_order_id, 
                               COALESCE(status, 'NEW') as status,
                               COALESCE(priority, 'MEDIUM') as priority,
                               latitude, longitude, accuracy, hidden
                        FROM meter_concern
                        WHERE 1=1
                    """
                    params = []
                    
                    if not show_hidden:
                        query += " AND (hidden = FALSE OR hidden IS NULL)"
                    
                    if search_term:
                        query += " AND (account_no ILIKE %s OR name ILIKE %s OR concern ILIKE %s)"
                        search_pattern = f"%{search_term}%"
                        params.extend([search_pattern, search_pattern, search_pattern])
                    
                    if status_filter_norm:
                        query += " AND UPPER(COALESCE(status, 'NEW')) = %s"
                        params.append(status_filter_norm)
                    
                    if priority_filter_norm:
                        query += " AND UPPER(COALESCE(priority, 'MEDIUM')) = %s"
                        params.append(priority_filter_norm)
                    
                    if type_filter_norm and type_filter_norm != 'BILLING':
                        continue
                    
                    if start_datetime:
                        query += " AND timestamp >= %s"
                        params.append(start_datetime)
                    
                    if end_datetime:
                        query += " AND timestamp <= %s"
                        params.append(end_datetime)
                    
                    query += " ORDER BY timestamp DESC"
                    
                    logger.debug(f"Meter Concern Query: {query}")
                    cur.execute(query, params)
                    rows = cur.fetchall()
                    logger.info(f"✅ Found {len(rows)} meter concern records")
                    
                    for row in rows:
                        complaints.append({
                            'time': row[7].strftime('%I:%M %p') if row[7] else 'N/A',
                            'customer_id': row[2] or 'N/A',
                            'job_order_id': row[8],
                            'issue_type': 'Billing',
                            'description': (row[6][:50] + '...') if row[6] and len(row[6]) > 50 else (row[6] or 'Meter concern'),
                            'priority': (row[10] or 'MEDIUM').upper(),
                            'status': (row[9] or 'NEW').upper(),
                            'source': 'Chatbot',
                            'table': 'meter_concern',
                            'record_id': row[0],
                            'customer_name': row[3] or 'Unknown',
                            'customer_phone': row[5] or 'N/A',
                            'full_data': {
                                'id': row[0],
                                'user_id': row[1],
                                'account_no': row[2],
                                'name': row[3],
                                'address': row[4],
                                'contact_number': row[5],
                                'concern': row[6],
                                'timestamp': row[7],
                                'job_order_id': row[8],
                                'status': row[9],
                                'priority': row[10],
                                'latitude': row[11],
                                'longitude': row[12],
                                'accuracy': row[13]
                            }
                        })
                
                elif table == 'power_outage_reports':
                    query = """
                        SELECT report_id, user_id, full_name, address, contact_number, 
                               job_order_id, 
                               COALESCE(status, 'NEW') as status,
                               COALESCE(issue_type, 'Power Outage') as issue_type,
                               COALESCE(priority, 'HIGH') as priority,
                               COALESCE(source, 'Outage Report') as source,
                               timestamp, latitude, longitude, accuracy, details, hidden
                        FROM power_outage_reports
                        WHERE 1=1
                    """
                    params = []
                    
                    if not show_hidden:
                        query += " AND (hidden = FALSE OR hidden IS NULL)"
                    
                    if search_term:
                        query += " AND (full_name ILIKE %s OR address ILIKE %s OR details ILIKE %s)"
                        search_pattern = f"%{search_term}%"
                        params.extend([search_pattern, search_pattern, search_pattern])
                    
                    if status_filter_norm:
                        query += " AND UPPER(COALESCE(status, 'NEW')) = %s"
                        params.append(status_filter_norm)
                    
                    if priority_filter_norm:
                        query += " AND UPPER(COALESCE(priority, 'HIGH')) = %s"
                        params.append(priority_filter_norm)
                    
                    if type_filter_norm and type_filter_norm not in ['POWER OUTAGE', 'POWER_OUTAGE']:
                        continue
                    
                    if start_datetime:
                        query += " AND timestamp >= %s"
                        params.append(start_datetime)
                    
                    if end_datetime:
                        query += " AND timestamp <= %s"
                        params.append(end_datetime)
                    
                    query += " ORDER BY timestamp DESC"
                    
                    logger.debug(f"Power Outage Query: {query}")
                    cur.execute(query, params)
                    rows = cur.fetchall()
                    logger.info(f"✅ Found {len(rows)} power outage records")
                    
                    for row in rows:
                        complaints.append({
                            'time': row[10].strftime('%I:%M %p') if row[10] else 'N/A',
                            'customer_id': row[1] or 'N/A',
                            'job_order_id': row[5],
                            'issue_type': (row[7] or 'Power Outage').strip().upper(),
                            'description': row[14] or 'Power outage reported',
                            'priority': (row[8] or 'HIGH').upper(),
                            'status': (row[6] or 'NEW').upper(),
                            'source': row[9] or 'Outage Report',
                            'table': 'power_outage_reports',
                            'record_id': row[0],
                            'customer_name': row[2] or 'Unknown',
                            'customer_phone': row[4] or 'N/A',
                            'full_data': {
                                'report_id': row[0],
                                'user_id': row[1],
                                'full_name': row[2],
                                'address': row[3],
                                'contact_number': row[4],
                                'job_order_id': row[5],
                                'status': row[6],
                                'issue_type': row[7],
                                'priority': row[8],
                                'source': row[9],
                                'timestamp': row[10],
                                'latitude': row[11],
                                'longitude': row[12],
                                'accuracy': row[13],
                                'details': row[14]
                            }
                        })
                        
            except Exception as table_error:
                logger.error(f"❌ Error processing table {table}: {table_error}")
                traceback.print_exc()
                continue

        # Sort complaints
        def sort_key(complaint):
            status_order = {'NEW': 0, 'ASSIGNED': 1, 'IN_PROGRESS': 2, 'RESOLVED': 3}
            priority_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}

            status = complaint.get('status', 'NEW').strip().upper()
            priority = complaint.get('priority', 'LOW').strip().upper()
            timestamp_value = complaint['full_data'].get('timestamp')

            timestamp_unix = 0
            if timestamp_value:
                if isinstance(timestamp_value, datetime):
                    timestamp_unix = timestamp_value.timestamp()
                else:
                    try:
                        timestamp_unix = datetime.strptime(str(timestamp_value), '%Y-%m-%d %H:%M:%S').timestamp()
                    except:
                        timestamp_unix = 0

            status_value = status_order.get(status, 4)
            priority_value = priority_order.get(priority, 4)

            return (status_value, priority_value, -timestamp_unix)

        complaints.sort(key=sort_key)
        logger.info(f"📊 Total complaints retrieved: {len(complaints)}")
        return complaints
        
    except Exception as e:
        logger.error(f"❌ Error getting unified complaints: {e}")
        traceback.print_exc()
        conn.rollback()
        return []
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

def calculate_chart_data(complaints):
    """Calculate chart data counts for different issue types"""
    technical_count = 0
    service_count = 0
    billing_count = 0
    power_outage_count = 0
    today_start = datetime.combine(date.today(), time(0, 0))
    today_end = datetime.combine(date.today(), time(23, 59, 59))
    for complaint in complaints:
        timestamp = complaint['full_data'].get('timestamp')
        issue_type = complaint.get('issue_type', '').strip().upper()
        if issue_type == 'POWER OUTAGE':
            technical_count += 1
            if timestamp and today_start <= timestamp <= today_end:
                power_outage_count += 1
        elif issue_type == 'SERVICE':
            service_count += 1
        elif issue_type == 'BILLING':
            billing_count += 1
    return technical_count, service_count, billing_count, power_outage_count

def extract_location_parts(location):
    """Extract barangay and town from location string"""
    if not location:
        return '', ''
    location = location.strip().lower()
    allowed_towns = {
        'tubungan': 'Tubungan', 'alimodian': 'Alimodian', 'cabatuan': 'Cabatuan',
        'guimbal': 'Guimbal', 'igbaras': 'Igbaras', 'leganes': 'Leganes', 'leon': 'Leon',
        'miagao': 'Miag-ao', 'miag-ao': 'Miag-ao', 'oton': 'Oton', 'pavia': 'Pavia',
        'san joaquin': 'San Joaquin', 'san miguel': 'San Miguel', 
        'sta barbara': 'Sta. Barbara', 'sta. barbara': 'Sta. Barbara', 'tigbauan': 'Tigbauan'
    }
    location = re.sub(r'\b(brgy|barangay)\.?\b', '', location, flags=re.IGNORECASE).strip()
    parts = [p.strip() for p in location.split(',') if p.strip()]
    brgy0 = parts[0] if parts else ''
    town0 = ''
    for key, proper_case in allowed_towns.items():
        if key in location:
            town0 = proper_case
            break
    brgy0 = brgy0.lstrip('.').title()
    return brgy0, town0

def notify_rasa(user_id, message):
    """Notify Rasa server of events"""
    try:
        if not RASA_URL:
            logger.warning("RASA_URL not set, skipping notification")
            return False
        response = requests.post(
            f"{RASA_URL}/webhooks/rest/webhook",
            json={"sender": user_id, "message": message},
            timeout=5
        )
        if response.status_code == 200:
            logger.info(f"Rasa notified successfully for user {user_id}")
            return True
        else:
            logger.error(f"Rasa notification failed: {response.status_code} - {response.text}")
            return False
    except Exception as e:
        logger.error(f"Error notifying Rasa: {e}")
        return False


def get_db_connection(db_config=DB_CONFIG):
    """Get a direct database connection"""
    try:
        conn = psycopg2.connect(**db_config)
        logger.debug(f"Database connection established to {db_config['database']}")
        return conn
    except Exception as e:
        logger.error(f"Failed to connect to database: {e}")
        return None

def release_connection(conn):
    """Close database connection"""
    if conn:
        try:
            conn.close()
            logger.debug("Database connection closed")
        except Exception as e:
            logger.error(f"Error closing connection: {e}")

def init_database():
    """Initialize database tables on startup"""
    conn = get_db_connection()
    if not conn:
        logger.error("Cannot initialize database - no connection")
        return False
    
    try:
        cur = conn.cursor()
        
        # Create power_outage_reports table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS power_outage_reports (
                report_id SERIAL PRIMARY KEY,
                full_name VARCHAR(255) NOT NULL,
                contact_number VARCHAR(20) NOT NULL,
                address TEXT NOT NULL,
                latitude DOUBLE PRECISION,
                longitude DOUBLE PRECISION,
                accuracy DOUBLE PRECISION,
                details TEXT NOT NULL,
                priority VARCHAR(20) DEFAULT 'HIGH',
                status VARCHAR(20) DEFAULT 'NEW',
                source VARCHAR(50) DEFAULT 'Web Form',
                job_order_id VARCHAR(50),
                hidden BOOLEAN DEFAULT FALSE,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        logger.info("power_outage_reports table initialized")
        
        # Create meter_concern table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS meter_concern (
                id SERIAL PRIMARY KEY,
                account_no VARCHAR(50) NOT NULL,
                name VARCHAR(255) NOT NULL,
                address TEXT NOT NULL,
                contact_number VARCHAR(20) NOT NULL,
                concern TEXT NOT NULL,
                priority VARCHAR(20) DEFAULT 'MEDIUM',
                status VARCHAR(20) DEFAULT 'NEW',
                source VARCHAR(50) DEFAULT 'Chatbot',
                job_order_id VARCHAR(50),
                latitude DOUBLE PRECISION,
                longitude DOUBLE PRECISION,
                accuracy DOUBLE PRECISION,
                hidden BOOLEAN DEFAULT FALSE,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        logger.info("meter_concern table initialized")
        
        # Create agent_queue table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS agent_queue (
                id SERIAL PRIMARY KEY,
                user_id VARCHAR(100) NOT NULL,
                full_name VARCHAR(255) NOT NULL,
                concern TEXT NOT NULL,
                contact_number VARCHAR(20) NOT NULL,
                priority VARCHAR(20) DEFAULT 'LOW',
                status VARCHAR(20) DEFAULT 'NEW',
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                hidden BOOLEAN DEFAULT FALSE,
                resumed BOOLEAN DEFAULT FALSE,
                latitude DOUBLE PRECISION,
                longitude DOUBLE PRECISION,
                accuracy DOUBLE PRECISION
            )
        """)
        logger.info("agent_queue table initialized")
        
        # Create indexes
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_power_timestamp 
            ON power_outage_reports(timestamp DESC);
            CREATE INDEX IF NOT EXISTS idx_power_status 
            ON power_outage_reports(status);
            CREATE INDEX IF NOT EXISTS idx_power_priority 
            ON power_outage_reports(priority);
        """)
        logger.info("Database indexes created")
        
        conn.commit()
        logger.info("Database initialization successful")
        return True
        
    except Exception as e:
        logger.error(f"Database initialization error: {e}")
        conn.rollback()
        return False
    finally:
        if 'cur' in locals():
            cur.close()
        release_connection(conn)

def init_joblist_database():
    """Initialize joblist database"""
    conn = get_db_connection(DB_JOBLIST)
    if not conn:
        logger.warning("Cannot connect to joblist database")
        return False
    
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS converted (
                unique_id TEXT PRIMARY KEY,
                creator TEXT,
                created TEXT,
                name TEXT,
                phone TEXT,
                address TEXT,
                latitude TEXT,
                longitude TEXT,
                town0 TEXT,
                brgy0 TEXT,
                priority TEXT,
                status TEXT DEFAULT 'pending',
                notes TEXT,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()
        logger.info("Joblist database initialized")
        return True
    except Exception as e:
        logger.error(f"Joblist database initialization error: {e}")
        return False
    finally:
        if 'cur' in locals():
            cur.close()
        release_connection(conn)

@app.route('/')
@login_required
def dashboard():
    """Render the dashboard with enhanced filtering - PROTECTED"""
    status_filter = request.args.get('status', 'All Status')
    priority_filter = request.args.get('priority', 'All Priorities')
    type_filter = request.args.get('type', 'All Types')
    search_term = request.args.get('search', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    time_from = request.args.get('time_from', '')
    time_to = request.args.get('time_to')
    
    # Get data
    queue = get_agent_queue()
    complaints = get_unified_complaints(
        status_filter, priority_filter, type_filter, search_term, False,
        date_from, date_to, time_from, time_to
    )
    
    # Calculate statistics
    technical_count, service_count, billing_count, power_outage_count = calculate_chart_data(complaints)
    
    critical_count = sum(1 for c in complaints if c.get('priority', '').upper() == 'CRITICAL')
    
    today_start = datetime.combine(date.today(), time(0, 0))
    today_end = datetime.combine(date.today(), time(23, 59, 59))
    resolved_today = sum(1 for c in complaints 
                        if c['status'] == 'RESOLVED' and 
                        c['full_data'].get('timestamp') and 
                        today_start <= c['full_data']['timestamp'] <= today_end)
    
    if queue:
        total_minutes = sum((datetime.now() - row.get('timestamp', datetime.now())).total_seconds() / 60 for row in queue)
        avg_wait_minutes = round(total_minutes / len(queue), 2)
    else:
        avg_wait_minutes = 0.0
    
    return render_template(
        'dashboard.html',
        queue=queue,
        complaints=complaints,
        current_status=status_filter,
        current_priority=priority_filter,
        current_type=type_filter,
        current_search=search_term,
        current_date_from=date_from,
        current_date_to=date_to,
        current_time_from=time_from,
        current_time_to=time_to,
        critical_count=critical_count,
        resolved_today=resolved_today,
        avg_wait_minutes=avg_wait_minutes,
        technical_count=technical_count,
        service_count=service_count,
        billing_count=billing_count,
        power_outage_count=power_outage_count,
        # Add user info to template
        current_user=session.get('username'),
        user_full_name=session.get('full_name'),
        user_role=session.get('role')
    )

# ============================================
#  ADMIN ROUTES (User Management)
# ============================================

@app.route('/admin/users')
@admin_required
def admin_users():
    """Admin page to manage users"""
    conn = get_rasa_connection()
    if not conn:
        flash('Database connection failed', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT id, username, email, full_name, role, is_active, created_at, last_login
            FROM users
            ORDER BY created_at DESC
        """)
        
        users = []
        for row in cur.fetchall():
            users.append({
                'id': row[0],
                'username': row[1],
                'email': row[2],
                'full_name': row[3],
                'role': row[4],
                'is_active': row[5],
                'created_at': row[6],
                'last_login': row[7]
            })
        
        return render_template('admin_users.html', users=users)
        
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

# Initialize users table on startup
create_users_table()

logger.info("=" * 60)
logger.info("🔐 Authentication System Initialized")
logger.info("=" * 60)
logger.info("Default Admin Credentials:")
logger.info("   Username: admin")
logger.info("   Password: admin123")
logger.info("=" * 60)
logger.info("⚠️  IMPORTANT: Change default password after first login!")
logger.info("=" * 60)

def get_priority_from_concern(concern: str) -> str:
    """Classify priority from concern text"""
    concern = (concern or "").lower()
    
    critical_keywords = [
        "fire", "explosion", "burning", "shock", "emergency", "danger",
        "hazard", "fallen wire", "live wire", "injured", "death"
    ]
    
    if any(word in concern for word in critical_keywords):
        return "CRITICAL"
    
    return "HIGH"

def generate_unique_id():
    """Generate unique ID for job orders"""
    return str(uuid.uuid4())[:12].upper()

def extract_location_parts(location):
    """Extract barangay and town from location"""
    if not location:
        return '', ''
    
    location = location.strip().lower()
    
    towns = {
        'tubungan': 'Tubungan', 'alimodian': 'Alimodian', 'cabatuan': 'Cabatuan',
        'guimbal': 'Guimbal', 'igbaras': 'Igbaras', 'leganes': 'Leganes',
        'leon': 'Leon', 'miagao': 'Miag-ao', 'oton': 'Oton', 'pavia': 'Pavia',
        'san joaquin': 'San Joaquin', 'san miguel': 'San Miguel',
        'sta barbara': 'Sta. Barbara', 'tigbauan': 'Tigbauan'
    }
    
    town = ''
    for key, value in towns.items():
        if key in location:
            town = value
            break
    
    parts = location.split(',')
    barangay = parts[0].strip() if parts else ''
    
    return barangay.title(), town

@app.route('/report')
def report():
    return render_template('report_outage.html')
# ============================================
#  1. EXPORT TO CSV
# ============================================

@app.route('/api/export_csv', methods=['GET'])
@login_required
def export_to_csv():
    """Export filtered complaints to CSV file"""
    try:
        # Get current filters from request
        status_filter = request.args.get('status', 'All Status')
        priority_filter = request.args.get('priority', 'All Priorities')
        type_filter = request.args.get('type', 'All Types')
        search_term = request.args.get('search', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        time_from = request.args.get('time_from', '')
        time_to = request.args.get('time_to', '')
        
        logger.info(f"📥 CSV Export requested by {session.get('username')}")
        
        # Get complaints with filters
        complaints = get_unified_complaints(
            status_filter, priority_filter, type_filter, search_term,
            False, date_from, date_to, time_from, time_to
        )
        
        # Create CSV in memory
        si = StringIO()
        writer = csv.writer(si)
        
        # Write headers
        headers = [
            'Complaint ID',
            'Date/Time',
            'Customer Name',
            'Customer ID',
            'Contact Number',
            'Issue Type',
            'Description',
            'Priority',
            'Status',
            'Job Order ID',
            'Source',
            'Address',
            'Latitude',
            'Longitude'
        ]
        writer.writerow(headers)
        
        # Write data rows
        for idx, c in enumerate(complaints, 1):
            row = [
                f"{c.get('table', 'UNK')}-{c.get('record_id', 'N/A')}",
                c.get('time', 'N/A'),
                c.get('customer_name', 'N/A'),
                c.get('customer_id', 'N/A'),
                c.get('customer_phone', 'N/A'),
                c.get('issue_type', 'N/A'),
                c.get('description', 'No description'),
                c.get('priority', 'N/A'),
                c.get('status', 'N/A'),
                c.get('job_order_id', 'Not assigned'),
                c.get('source', 'N/A'),
                c['full_data'].get('address', 'N/A') if 'full_data' in c else 'N/A',
                c['full_data'].get('latitude', 'N/A') if 'full_data' in c else 'N/A',
                c['full_data'].get('longitude', 'N/A') if 'full_data' in c else 'N/A'
            ]
            writer.writerow(row)
        
        # Create filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"ILECO1_Complaints_{timestamp}.csv"
        
        # Create response
        output = BytesIO()
        output.write(si.getvalue().encode('utf-8-sig'))  # UTF-8 with BOM for Excel
        output.seek(0)
        
        # Log the export
        log_audit(
            session.get('user_id'),
            'EXPORT_CSV',
            f'Exported {len(complaints)} complaints to CSV',
            request.remote_addr
        )
        
        logger.info(f"✅ CSV exported: {len(complaints)} complaints")
        
        return send_file(
            output,
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"❌ CSV export error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============================================
#  2. EXPORT TO EXCEL (ENHANCED)
# ============================================

@app.route('/api/export_excel', methods=['GET'])
@login_required
def export_to_excel():
    """Export filtered complaints to Excel with formatting"""
    try:
        # Get current filters
        status_filter = request.args.get('status', 'All Status')
        priority_filter = request.args.get('priority', 'All Priorities')
        type_filter = request.args.get('type', 'All Types')
        search_term = request.args.get('search', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        time_from = request.args.get('time_from', '')
        time_to = request.args.get('time_to', '')
        
        logger.info(f"📥 Excel Export requested by {session.get('username')}")
        
        # Get complaints with filters
        complaints = get_unified_complaints(
            status_filter, priority_filter, type_filter, search_term,
            False, date_from, date_to, time_from, time_to
        )
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Complaints Report"
        
        # Define styles
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Priority color mapping
        priority_colors = {
            'CRITICAL': 'B71C1C',
            'HIGH': 'F44336',
            'MEDIUM': 'FFC107',
            'LOW': '4CAF50'
        }
        
        # Status color mapping
        status_colors = {
            'NEW': 'CCE5FF',
            'ASSIGNED': 'FFF3CD',
            'IN_PROGRESS': 'D1ECF1',
            'RESOLVED': 'D4EDDA'
        }
        
        # Add title row
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = 'ILECO-1 COMPLAINT MANAGEMENT REPORT'
        title_cell.font = Font(bold=True, size=16, color="003366")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add metadata
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
        ws['A3'] = f"Generated by: {session.get('full_name', session.get('username'))}"
        ws['A4'] = f"Total Complaints: {len(complaints)}"
        
        # Add filters info
        filter_row = 5
        if status_filter != 'All Status':
            ws[f'A{filter_row}'] = f"Status Filter: {status_filter}"
            filter_row += 1
        if priority_filter != 'All Priorities':
            ws[f'A{filter_row}'] = f"Priority Filter: {priority_filter}"
            filter_row += 1
        if date_from or date_to:
            ws[f'A{filter_row}'] = f"Date Range: {date_from or 'Any'} to {date_to or 'Any'}"
            filter_row += 1
        
        # Headers start row
        header_row = filter_row + 1
        
        # Write headers
        headers = [
            'ID', 'Date/Time', 'Customer Name', 'Customer ID', 'Contact',
            'Issue Type', 'Description', 'Priority', 'Status', 'Job Order',
            'Source', 'Address', 'Latitude', 'Longitude'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        data_start_row = header_row + 1
        for idx, complaint in enumerate(complaints, 1):
            row_num = data_start_row + idx - 1
            
            # Data values
            data = [
                f"{complaint.get('table', 'UNK')}-{complaint.get('record_id', 'N/A')}",
                complaint.get('time', 'N/A'),
                complaint.get('customer_name', 'N/A'),
                complaint.get('customer_id', 'N/A'),
                complaint.get('customer_phone', 'N/A'),
                complaint.get('issue_type', 'N/A'),
                complaint.get('description', 'No description')[:100],  # Truncate long descriptions
                complaint.get('priority', 'N/A'),
                complaint.get('status', 'N/A'),
                complaint.get('job_order_id', 'Not assigned'),
                complaint.get('source', 'N/A'),
                complaint['full_data'].get('address', 'N/A') if 'full_data' in complaint else 'N/A',
                complaint['full_data'].get('latitude', 'N/A') if 'full_data' in complaint else 'N/A',
                complaint['full_data'].get('longitude', 'N/A') if 'full_data' in complaint else 'N/A'
            ]
            
            # Write data
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                
                # Apply priority color
                if col_num == 8:  # Priority column
                    priority = complaint.get('priority', '').upper()
                    if priority in priority_colors:
                        cell.fill = PatternFill(start_color=priority_colors[priority], 
                                              end_color=priority_colors[priority], 
                                              fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                
                # Apply status color
                if col_num == 9:  # Status column
                    status = complaint.get('status', '').upper()
                    if status in status_colors:
                        cell.fill = PatternFill(start_color=status_colors[status], 
                                              end_color=status_colors[status], 
                                              fill_type="solid")
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze panes (freeze header)
        ws.freeze_panes = f'A{data_start_row}'
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws['A1'] = 'COMPLAINT SUMMARY'
        summary_ws['A1'].font = Font(bold=True, size=14)
        
        # Calculate statistics
        total = len(complaints)
        critical = sum(1 for c in complaints if c.get('priority', '').upper() == 'CRITICAL')
        high = sum(1 for c in complaints if c.get('priority', '').upper() == 'HIGH')
        new = sum(1 for c in complaints if c.get('status', '').upper() == 'NEW')
        resolved = sum(1 for c in complaints if c.get('status', '').upper() == 'RESOLVED')
        
        summary_data = [
            ['', ''],
            ['Total Complaints', total],
            ['', ''],
            ['BY PRIORITY', ''],
            ['Critical', critical],
            ['High', high],
            ['', ''],
            ['BY STATUS', ''],
            ['New', new],
            ['Resolved', resolved],
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 2):
            summary_ws[f'A{row_idx}'] = label
            summary_ws[f'B{row_idx}'] = value
            if label and 'BY' in label:
                summary_ws[f'A{row_idx}'].font = Font(bold=True)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"ILECO1_Complaints_{timestamp}.xlsx"
        
        # Log the export
        log_audit(
            session.get('user_id'),
            'EXPORT_EXCEL',
            f'Exported {len(complaints)} complaints to Excel',
            request.remote_addr
        )
        
        logger.info(f"✅ Excel exported: {len(complaints)} complaints")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"❌ Excel export error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500




# NEW: Enhanced complaints_with_location using PostGIS spatial queries
@app.route('/api/complaints_with_location', methods=['GET'])
def get_complaints_with_location():
    """
    Fetch complaints with location data using PostGIS geometry
    Much faster than lat/lng due to spatial indexing!
    """
    issue_type = request.args.get('type', 'All Types')
    search = request.args.get('search', '')
    status = request.args.get('status', 'All Status')
    priority = request.args.get('priority', 'All Priorities')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    time_from = request.args.get('time_from')
    time_to = request.args.get('time_to')
    
    status_norm = status.strip().upper() if status and status != 'All Status' else None
    priority_norm = priority.strip().upper() if priority and priority != 'All Priorities' else None
    type_norm = issue_type.strip().upper() if issue_type and issue_type != 'All Types' else None
    
    logger.info(f"🗺️ Fetching complaints with PostGIS geometry - Filters: {status_norm}, {priority_norm}, {type_norm}")
    
    start_datetime, end_datetime = parse_datetime_filters(date_from, date_to, time_from, time_to)
    conn = get_rasa_connection()
    if not conn:
        logger.error("Database connection failed")
        return jsonify({'success': False, 'error': 'Database connection failed', 'complaints': []}), 500
    
    complaints = []
    try:
        cur = conn.cursor()
        
        # Power outage reports - Using PostGIS geometry!
        query = """
            SELECT 
                report_id,
                full_name AS customer,
                'N/A' AS customer_id,
                address,
                ST_Y(geom) AS lat,
                ST_X(geom) AS lng,
                COALESCE(issue_type, 'Power Outage') AS issue_type,
                details AS description,
                COALESCE(priority, 'HIGH') AS priority,
                COALESCE(status, 'NEW') AS status,
                TO_CHAR(timestamp, 'HH12:MI AM') AS time,
                contact_number AS contact,
                'Power Outage Report' AS source,
                COALESCE(accuracy, 0) AS accuracy,
                details,
                job_order_id,
                ST_AsText(geom) AS geom_wkt
            FROM power_outage_reports
            WHERE geom IS NOT NULL
              AND (hidden = FALSE OR hidden IS NULL)
        """
        params = []
        
        if type_norm and type_norm in ['POWER OUTAGE', 'POWER_OUTAGE']:
            query += " AND UPPER(REPLACE(issue_type, ' ', '_')) = 'POWER_OUTAGE'"
        if search:
            query += " AND (full_name ILIKE %s OR address ILIKE %s)"
            params.extend([f"%{search}%", f"%{search}%"])
        if status_norm:
            query += " AND UPPER(COALESCE(status, 'NEW')) = %s"
            params.append(status_norm)
        if priority_norm:
            query += " AND UPPER(COALESCE(priority, 'HIGH')) = %s"
            params.append(priority_norm)
        if start_datetime:
            query += " AND timestamp >= %s"
            params.append(start_datetime)
        if end_datetime:
            query += " AND timestamp <= %s"
            params.append(end_datetime)
        
        cur.execute(query, params)
        power_outages = cur.fetchall()
        
        for row in power_outages:
            complaints.append({
                'id': f"PO-{row[0]:06d}",
                'customer': row[1] or 'Unknown',
                'customerId': row[2],
                'address': row[3] or 'N/A',
                'lat': float(row[4]) if row[4] else None,
                'lng': float(row[5]) if row[5] else None,
                'issueType': row[6] or 'Power Outage',
                'description': row[7] or 'Power outage reported',
                'priority': (row[8] or 'HIGH').upper(),
                'status': (row[9] or 'NEW').upper(),
                'time': row[10] or 'N/A',
                'contact': row[11] or 'N/A',
                'source': row[12],
                'accuracy': row[13] or 0,
                'details': row[14] or 'No details provided',
                'job_order_id': row[15],
                'geometry': row[16]  # NEW: Include WKT geometry
            })
        
        logger.info(f"✅ Retrieved {len(complaints)} complaints using PostGIS spatial indexing")
        
        return jsonify({
            'success': True,
            'complaints': complaints,
            'count': len(complaints),
            'postgis_enabled': True
        })
        
    except Exception as e:
        logger.error(f"Database error: {e}")
        traceback.print_exc()
        conn.rollback()
        return jsonify({'success': False, 'error': str(e), 'complaints': []}), 500
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

@app.route('/api/geocode', methods=['POST'])
@limiter.limit("10 per minute")
def geocode_address():
    """Geocode an address to coordinates using Google Maps API or Nominatim"""
    try:
        data = request.get_json()
        address = data.get('address', '').strip()
        if not address:
            logger.warning("Address missing in geocode request")
            return jsonify({'error': 'Address is required'}), 400
        headers = {'User-Agent': 'ILECO-Chatbot/1.0'}
        if GOOGLE_API_KEY:
            google_url = "https://maps.googleapis.com/maps/api/geocode/json"
            params = {
                'address': f"{address}, Iloilo, Philippines",
                'key': GOOGLE_API_KEY
            }
            response = requests.get(google_url, params=params, headers=headers, timeout=5)
            data = response.json()
            if data.get('status') == 'OK' and data.get('results'):
                location = data['results'][0]['geometry']['location']
                logger.info(f"Geocoded address {address} using Google Maps")
                return jsonify({
                    'success': True,
                    'lat': float(location['lat']),
                    'lng': float(location['lng']),
                    'display_name': data['results'][0].get('formatted_address', address)
                })
            elif data.get('status') == 'OVER_QUERY_LIMIT':
                logger.warning("Google API quota exceeded")
                return jsonify({'error': 'Google API quota exceeded'}), 429
        nominatim_url = "https://nominatim.openstreetmap.org/search"
        params = {
            'q': f"{address}, Iloilo, Philippines",
            'format': 'json',
            'limit': 1
        }
        response = requests.get(nominatim_url, params=params, headers=headers, timeout=5)
        data = response.json()
        if data:
            logger.info(f"Geocoded address {address} using Nominatim")
            return jsonify({
                'success': True,
                'lat': float(data[0]['lat']),
                'lng': float(data[0]['lon']),
                'display_name': data[0].get('display_name', address)
            })
        logger.warning(f"Could not geocode address {address}, using default location")
        return jsonify({
            'success': False,
            'lat': 11.5854,
            'lng': 122.7507,
            'message': 'Could not geocode address, using default location'
        })
    except Exception as e:
        logger.error(f"Geocoding error: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'lat': 11.5854,
            'lng': 122.7507
        }), 500

@app.route('/api/update_location/<table>/<int:record_id>', methods=['POST'])
def update_complaint_location(table, record_id):
    """Update location for a complaint record"""
    try:
        data = request.get_json()
        latitude = data.get('latitude')
        longitude = data.get('longitude')
        accuracy = data.get('accuracy')
        if latitude is None or longitude is None:
            logger.warning("Missing latitude or longitude in update_location")
            return jsonify({'error': 'Latitude and longitude are required'}), 400
        try:
            latitude = float(latitude)
            longitude = float(longitude)
            accuracy = float(accuracy) if accuracy is not None else None
        except (ValueError, TypeError):
            logger.warning("Invalid latitude, longitude, or accuracy format")
            return jsonify({'error': 'Invalid latitude, longitude, or accuracy format'}), 400
        allowed_tables = ['power_outage_reports', 'meter_concern', 'agent_queue']
        if table not in allowed_tables:
            logger.warning(f"Invalid table name: {table}")
            return jsonify({'error': 'Invalid table name'}), 400
        conn = get_rasa_connection()
        if not conn:
            logger.error("Database connection failed for update_location")
            return jsonify({'error': 'Database connection failed'}), 500
        try:
            cur = conn.cursor()
            pk_column = 'report_id' if table == 'power_outage_reports' else 'id'
            cur.execute(
                f"UPDATE {table} SET latitude = %s, longitude = %s, accuracy = %s WHERE {pk_column} = %s",
                (latitude, longitude, accuracy, record_id)
            )
            if cur.rowcount == 0:
                logger.warning(f"Record not found: {table}/{record_id}")
                return jsonify({'error': 'Record not found'}), 404
            conn.commit()
            logger.info(f"Location updated for {table}/{record_id}")
            return jsonify({
                'success': True,
                'message': 'Location updated successfully',
                'latitude': latitude,
                'longitude': longitude,
                'accuracy': accuracy
            })
        except Exception as e:
            logger.error(f"Error updating location: {e}")
            conn.rollback()
            return jsonify({'error': str(e)}), 500
        finally:
            if 'cur' in locals():
                cur.close()
            release_db_connection(conn, db_pool)
    except Exception as e:
        logger.error(f"Error in update_complaint_location: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/assign_job_order/<table>/<int:record_id>', methods=['POST'])
def assign_job_order(table, record_id):
    """Assign a job order to the converted table in joblist database"""
    try:
        ensure_converted_table_joblist()
        conn = get_rasa_connection()
        if not conn:
            logger.error("Database connection failed for assign_job_order")
            return jsonify({'success': False, 'error': 'Database connection failed'})

        try:
            cur = conn.cursor()

            # Fetch record data based on table type
            if table == 'power_outage_reports':
                # ✅ CRITICAL: Fetch ALL fields including incident details
                cur.execute("""
                    SELECT 
                        report_id, user_id, full_name, address, contact_number, job_order_id,
                        COALESCE(status, 'NEW') as status,
                        COALESCE(issue_type, 'Power Outage') as issue_type,
                        COALESCE(priority, 'HIGH') as priority,
                        COALESCE(source, 'Outage Report') as source,
                        timestamp, latitude, longitude, accuracy, details, hidden,
                        email, account_number,
                        incident_type_detail, affected_area, incident_time, duration, landmark
                    FROM power_outage_reports 
                    WHERE report_id = %s
                """, (record_id,))

                row = cur.fetchone()

                if row:
                    timestamp = row[10]
                    if isinstance(timestamp, str):
                        try:
                            timestamp = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            timestamp = datetime.now()

                    # ✅ Map incident types to readable names
                    incident_type_map = {
                        'power_outage': 'Power Outage / No Electricity',
                        'fallen_wire': 'Fallen Wire / Pole Down',
                        'transformer_issue': 'Transformer Problem / Explosion',
                        'fire_hazard': 'Fire / Burning Smell',
                        'sparking': 'Sparking / Arcing',
                        'voltage_issue': 'Voltage Problem (High/Low)',
                        'equipment_damage': 'Equipment Damage',
                        'other': 'Other Emergency'
                    }

                    affected_area_map = {
                        'single_house': 'Single House/Building',
                        'multiple_houses': 'Multiple Houses (2-5)',
                        'street': 'Entire Street/Purok',
                        'barangay': 'Entire Barangay',
                        'multiple_barangay': 'Multiple Barangays',
                        'unknown': 'Unknown'
                    }

                    duration_map = {
                        'just_now': 'Just Now (Less than 15 min)',
                        '15_30min': '15-30 minutes',
                        '30_60min': '30 minutes - 1 hour',
                        '1_3hours': '1-3 hours',
                        '3_6hours': '3-6 hours',
                        'more_6hours': 'More than 6 hours'
                    }

                    # ✅ FIX: Get raw values with proper error handling and conversion to string
                    incident_type_raw = str(row[17]) if row[17] else 'power_outage'
                    affected_area_raw = str(row[18]) if row[18] else 'unknown'
                    incident_time_raw = str(row[19]) if row[19] else None
                    duration_raw = str(row[20]) if row[20] else 'just_now'
                    landmark_raw = str(row[21]) if row[21] else None

            elif table == 'meter_concern':
                cur.execute("""
                    SELECT id, user_id, account_no, name, address, contact_number, concern, 
                           timestamp, job_order_id, status, latitude, longitude, accuracy
                    FROM meter_concern 
                    WHERE id = %s
                """, (record_id,))
                db_row = cur.fetchone()

                if not db_row:
                    logger.warning(f"Record not found: meter_concern/{record_id}")
                    return jsonify({'success': False, 'error': 'Record not found'})

                full_name = db_row[3] or 'Unknown'
                phone = db_row[5] or 'N/A'
                location_raw = db_row[4] or 'N/A'
                concern = db_row[6] or 'Meter concern'
                priority = 'MEDIUM'
                existing_job_order = db_row[8] or ''
                latitude = db_row[10] or ''
                longitude = db_row[11] or ''
                accuracy = db_row[12] or ''

            elif table == 'agent_queue':
                cur.execute("""
                    SELECT id, user_id, full_name, concern, contact_number, priority, timestamp, 
                           job_order_id, latitude, longitude, accuracy
                    FROM agent_queue 
                    WHERE id = %s
                """, (record_id,))
                db_row = cur.fetchone()

                if not db_row:
                    logger.warning(f"Record not found: agent_queue/{record_id}")
                    return jsonify({'success': False, 'error': 'Record not found'})

                full_name = db_row[2] or 'Unknown'
                phone = db_row[4] or 'N/A'
                location_raw = ''
                concern = db_row[3] or 'Service request'
                priority = db_row[5] or 'LOW'
                existing_job_order = db_row[7] or ''
                latitude = db_row[8] or ''
                longitude = db_row[9] or ''
                accuracy = db_row[10] or ''

            else:
                logger.warning(f"Invalid table name: {table}")
                return jsonify({'success': False, 'error': 'Invalid table name'})

            # Parse location
            brgy0, town0 = extract_location_parts(location_raw)

            # Generate unique ID for job order
            converted_unique_id = str(uuid.uuid4().int)[:10]

            # Prepare converted data
            converted_data = {
                'unique_id': converted_unique_id,
                'creator': phone or full_name or 'Dashboard',
                'created': datetime.now().strftime('%m/%d/%y %I:%M:%S %p'),
                'follower': phone or full_name or 'Unknown',
                'followed': datetime.now().strftime('%m/%d/%y %I:%M:%S %p'),
                'name': full_name,
                'spinners': phone or '',
                'town0': town0,
                'brgy0': brgy0,
                'town': 'Select Town',
                'brgy': 'Select Brgy',
                'town2': 'Select Town',
                'brgy2': 'Select Brgy',
                'assignedto': town0 or '',
                'status': 'Select Status',
                'subs': 'Substation',
                'feeder': 'Feeder',
                'section': 'Category',
                'cause': concern,
                'equip': 'Equipment',
                'type': 'high' if any(word in concern.lower() for word in ['power', 'outage', 'emergency']) else 'low',
                'notes': (
                    f"Original Job Order: {existing_job_order} | Priority: {priority} | {concern}"
                    if existing_job_order else f"Priority: {priority} | {concern}"
                ),
                'landmark': '',
                'phone': phone,
                'location': location_raw,
                'latitude': str(latitude),
                'longitude': str(longitude),
                'actiontaken': 'Pending'
            }

            # Insert into joblist database
            joblist_conn = get_joblist_connection()
            if not joblist_conn:
                logger.error("Joblist database connection failed")
                return jsonify({'success': False, 'error': 'Joblist database connection failed'})

            try:
                joblist_cur = joblist_conn.cursor()
                columns = list(converted_data.keys())
                values = list(converted_data.values())
                placeholders = ', '.join(['%s'] * len(values))
                insert_query = f"INSERT INTO converted ({', '.join(columns)}) VALUES ({placeholders})"
                joblist_cur.execute(insert_query, values)
                joblist_conn.commit()
                logger.info(f"Job order {converted_unique_id} created in joblist database")

            except Exception as e:
                logger.error(f"Error inserting into converted table: {e}")
                joblist_conn.rollback()
                return jsonify({'success': False, 'error': str(e)})

            finally:
                if 'joblist_cur' in locals():
                    joblist_cur.close()
                release_db_connection(joblist_conn, joblist_db_pool)

            # ✅ FIX: Update status to ASSIGNED and set job_order_id in the original table
            pk_column = 'report_id' if table == 'power_outage_reports' else 'id'

            # Update both status AND job_order_id in a single query
            cur.execute(
                f"UPDATE {table} SET status = %s, job_order_id = %s WHERE {pk_column} = %s",
                ('ASSIGNED', converted_unique_id, record_id)
            )

            if cur.rowcount == 0:
                logger.warning(f"No rows updated for {table}/{record_id}")
                conn.rollback()
                return jsonify({'success': False, 'error': 'Failed to update record status'})

            # ✅ CRITICAL: Commit the changes!
            conn.commit()
            logger.info(f"Status updated to ASSIGNED for {table}/{record_id}")

            return jsonify({
                'success': True,
                'message': 'Job order assigned successfully',
                'converted_unique_id': converted_unique_id,
                'original_job_order_id': existing_job_order,
                'customer_name': full_name,
                'location': location_raw,
                'brgy0': brgy0,
                'town0': town0,
                'new_status': 'ASSIGNED'
            })

        except Exception as e:
            logger.error(f"Error assigning job order: {e}")
            traceback.print_exc()
            conn.rollback()
            return jsonify({'success': False, 'error': str(e)})

        finally:
            if 'cur' in locals():
                cur.close()
            release_db_connection(conn, db_pool)

    except Exception as e:
        logger.error(f"Unexpected error in assign_job_order: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/view/<table>/<int:record_id>', methods=['GET'])  # ✅ Changed from /view_complaint/ to /view/
def view_complaint(table, record_id):
    """View detailed information about a specific complaint"""
    try:
        conn = get_rasa_connection()
        if not conn:
            logger.error("Database connection failed for view_complaint")
            return jsonify({'error': 'Database connection failed'}), 500
        
        try:
            cur = conn.cursor()
            complaint_data = {}
            
            if table == 'power_outage_reports':
                cur.execute("""
                    SELECT 
                        report_id, user_id, full_name, address, contact_number, job_order_id,
                        COALESCE(status, 'NEW') as status,
                        COALESCE(issue_type, 'Power Outage') as issue_type,
                        COALESCE(priority, 'HIGH') as priority,
                        COALESCE(source, 'Outage Report') as source,
                        timestamp, latitude, longitude, accuracy, details, hidden,
                        email, account_number,
                        incident_type_detail, affected_area, incident_time, duration, landmark
                    FROM power_outage_reports 
                    WHERE report_id = %s
                """, (record_id,))
                
                row = cur.fetchone()
                
                if row:
                    timestamp = row[10]
                    if isinstance(timestamp, str):
                        try:
                            timestamp = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            timestamp = datetime.now()
                    
                    incident_type_map = {
                        'power_outage': 'Power Outage / No Electricity',
                        'fallen_wire': 'Fallen Wire / Pole Down',
                        'transformer_issue': 'Transformer Problem / Explosion',
                        'fire_hazard': 'Fire / Burning Smell',
                        'sparking': 'Sparking / Arcing',
                        'voltage_issue': 'Voltage Problem (High/Low)',
                        'equipment_damage': 'Equipment Damage',
                        'other': 'Other Emergency'
                    }
                    
                    affected_area_map = {
                        'single_house': 'Single House/Building',
                        'multiple_houses': 'Multiple Houses (2-5)',
                        'street': 'Entire Street/Purok',
                        'barangay': 'Entire Barangay',
                        'multiple_barangay': 'Multiple Barangays',
                        'unknown': 'Unknown'
                    }
                    
                    duration_map = {
                        'just_now': 'Just Now (Less than 15 min)',
                        '15_30min': '15-30 minutes',
                        '30_60min': '30 minutes - 1 hour',
                        '1_3hours': '1-3 hours',
                        '3_6hours': '3-6 hours',
                        'more_6hours': 'More than 6 hours'
                    }
                    
                    incident_type_raw = str(row[17]) if row[17] else 'power_outage'
                    affected_area_raw = str(row[18]) if row[18] else 'unknown'
                    incident_time_raw = str(row[19]) if row[19] else None
                    duration_raw = str(row[20]) if row[20] else 'just_now'
                    landmark_raw = str(row[21]) if row[21] else None
                    
                    complaint_data = {
                        'title': 'Power Outage Report Details',
                        'table_type': 'Power Outage',
                        'priority_level': row[8] or 'HIGH',
                        'status_level': row[6] or 'NEW',
                        'details': {
                            'Complaint ID': f"PO-{row[0]:06d}",
                            'Customer Name': row[2] or 'N/A',
                            'Email': row[15] or 'N/A',
                            'Account Number': row[16] or 'N/A',
                            'Contact Information': row[4] or 'N/A',
                            'Affected Location': row[3] or 'N/A',
                            'Nearest Landmark': landmark_raw or 'Not specified',
                            
                            '--- INCIDENT DETAILS ---': '',
                            '🔴 Type of Incident': incident_type_map.get(incident_type_raw, incident_type_raw.replace('_', ' ').title()),
                            '🏘️ Affected Area': affected_area_map.get(affected_area_raw, affected_area_raw.replace('_', ' ').title()),
                            '⏰ Incident Start Time': incident_time_raw or 'Not recorded',
                            '⏱️ Duration': duration_map.get(duration_raw, duration_raw.replace('_', ' ').title()),
                            
                            '--- GENERAL INFORMATION ---': '',
                            'Issue Type': row[7] or 'Power Outage',
                            'Priority': row[8] or 'HIGH',
                            'Source': row[9] or 'Outage Report',
                            'Status': row[6] or 'NEW',
                            'Job Order ID': row[5] or 'Not assigned yet',
                            
                            '--- TIMING ---': '',
                            'Date Reported': timestamp.strftime('%B %d, %Y') if timestamp else 'N/A',
                            'Time Reported': timestamp.strftime('%I:%M %p') if timestamp else 'N/A',
                            
                            '--- LOCATION DATA ---': '',
                            'Latitude': row[11] or 'N/A',
                            'Longitude': row[12] or 'N/A',
                            'GPS Accuracy': f"{row[13]}m" if row[13] else 'N/A',
                            
                            '--- DETAILED DESCRIPTION ---': '',
                            'Details': row[14] or 'No details provided'
                        }
                    }
                    
            elif table == 'meter_concern':
                cur.execute("""
                    SELECT id, user_id, account_no, name, address, contact_number, concern, 
                           timestamp, job_order_id, COALESCE(status, 'NEW') as status, 
                           latitude, longitude, accuracy
                    FROM meter_concern WHERE id = %s
                """, (record_id,))
                row = cur.fetchone()
                
                if row:
                    timestamp = row[7]
                    if isinstance(timestamp, str):
                        try:
                            timestamp = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            timestamp = datetime.now()
                    
                    complaint_data = {
                        'title': 'Billing Concern Details',
                        'table_type': 'Billing Issue',
                        'priority_level': 'MEDIUM',
                        'status_level': row[9] or 'NEW',
                        'details': {
                            'Complaint ID': f"BC-{row[0]:06d}",
                            'Account Number': row[2] or 'N/A',
                            'Customer Name': row[3] or 'N/A',
                            'Service Address': row[4] or 'N/A',
                            'Contact Number': row[5] or 'N/A',
                            'Concern Details': row[6] or 'No concern details provided',
                            'Status': row[9] or 'NEW',
                            'Job Order ID': row[8] or 'Not yet assigned',
                            'Date Reported': timestamp.strftime('%B %d, %Y') if timestamp else 'N/A',
                            'Time Reported': timestamp.strftime('%I:%M %p') if timestamp else 'N/A',
                            'Latitude': row[10] or 'N/A',
                            'Longitude': row[11] or 'N/A',
                            'GPS Accuracy': f"{row[12]}m" if row[12] else 'N/A'
                        }
                    }
                    
            elif table == 'agent_queue':
                cur.execute("""
                    SELECT id, user_id, full_name, concern, contact_number, priority, timestamp, 
                           COALESCE(status, 'NEW') as status, latitude, longitude, accuracy
                    FROM agent_queue WHERE id = %s
                """, (record_id,))
                row = cur.fetchone()
                
                if row:
                    timestamp = row[6]
                    if isinstance(timestamp, str):
                        try:
                            timestamp = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            timestamp = datetime.now()
                    
                    complaint_data = {
                        'title': 'Service Request Details',
                        'table_type': 'Service Request',
                        'priority_level': row[5] or 'LOW',
                        'status_level': row[7] or 'NEW',
                        'details': {
                            'Complaint ID': f"SR-{row[0]:06d}",
                            'Customer ID': row[1] or 'N/A',
                            'Full Name': row[2] or 'N/A',
                            'Contact Number': row[4] or 'N/A',
                            'Concern Description': row[3] or 'No description provided',
                            'Priority Level': row[5] or 'LOW',
                            'Status': row[7] or 'NEW',
                            'Date Submitted': timestamp.strftime('%B %d, %Y') if timestamp else 'N/A',
                            'Time Submitted': timestamp.strftime('%I:%M %p') if timestamp else 'N/A',
                            'Request Type': 'Agent Assistance Request',
                            'Latitude': row[8] or 'N/A',
                            'Longitude': row[9] or 'N/A',
                            'GPS Accuracy': f"{row[10]}m" if row[10] else 'N/A'
                        }
                    }
            
            if not complaint_data:
                return jsonify({'error': 'Complaint not found'}), 404
            
            logger.info(f"Retrieved complaint details for {table}/{record_id}")
            return jsonify(complaint_data)
            
        except Exception as e:
            logger.error(f"Error viewing complaint: {e}")
            traceback.print_exc()
            if conn:
                conn.rollback()
            return jsonify({'error': str(e)}), 500
        finally:
            if 'cur' in locals():
                cur.close()
            release_db_connection(conn, db_pool)
            
    except Exception as e:
        logger.error(f"Error in view_complaint: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/resume_conversation/<int:row_id>', methods=['POST'])
def resume_conversation(row_id):
    """Resume conversation with user from agent queue"""
    conn = get_rasa_connection()
    if not conn:
        logger.error("Database connection failed for resume_conversation")
        return jsonify({'success': False, 'error': 'Database connection failed'})
    try:
        cur = conn.cursor()
        cur.execute("SELECT user_id FROM agent_queue WHERE id = %s", (row_id,))
        result = cur.fetchone()
        user_id_value = result[0] if result else None
        if user_id_value:
            response = requests.post(
                f"{RASA_URL}/conversations/{user_id_value}/execute",
                json={"name": "action_resume_conversation"},
                timeout=5
            )
            if response.status_code != 200:
                logger.error(f"Failed to resume Rasa conversation: {response.text}")
            else:
                logger.info("Conversation resumed successfully")
            ensure_columns(conn, 'agent_queue')
            cur.execute("UPDATE agent_queue SET resumed = TRUE WHERE id = %s", (row_id,))
            conn.commit()
        return redirect('/')
    except Exception as e:
        logger.error(f"Error resuming conversation: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)


@app.route('/delete_complaint/<table>/<record_id>', methods=['POST'])
def delete_complaint(table, record_id):
    """Delete (hide) a complaint from the dashboard"""
    try:
        # Convert record_id to int
        try:
            record_id = int(record_id)
        except ValueError:
            logger.error(f"Invalid record_id: {record_id}")
            return jsonify({'success': False, 'error': 'Invalid record ID'}), 400
        
        # Validate table name
        allowed_tables = ['power_outage_reports', 'meter_concern', 'agent_queue']
        if table not in allowed_tables:
            logger.warning(f"Invalid table name: {table}")
            return jsonify({'success': False, 'error': 'Invalid table name'}), 400
        
        conn = get_rasa_connection()
        if not conn:
            logger.error("Database connection failed for delete_complaint")
            return jsonify({'success': False, 'error': 'Database connection failed'}), 500
        
        try:
            cur = conn.cursor()
            
            # Ensure the hidden column exists
            ensure_columns(conn, table)
            
            # Get the primary key column name
            pk_column = 'report_id' if table == 'power_outage_reports' else 'id'
            
            # Check if record exists
            cur.execute(f"SELECT {pk_column} FROM {table} WHERE {pk_column} = %s", (record_id,))
            if not cur.fetchone():
                logger.warning(f"Record not found for deletion: {table}/{record_id}")
                return jsonify({'success': False, 'error': 'Record not found'}), 404
            
            # Set hidden = TRUE (soft delete)
            cur.execute(f"UPDATE {table} SET hidden = TRUE WHERE {pk_column} = %s", (record_id,))
            
            if cur.rowcount == 0:
                logger.warning(f"No rows updated for {table}/{record_id}")
                conn.rollback()
                return jsonify({'success': False, 'error': 'Failed to delete record'}), 500
            
            conn.commit()
            
            logger.info(f"✅ Complaint deleted (hidden): {table}/{record_id}")
            
            return jsonify({
                'success': True,
                'message': 'Complaint removed successfully',
                'table': table,
                'record_id': record_id
            })
            
        except Exception as e:
            logger.error(f"Error deleting complaint: {e}")
            traceback.print_exc()
            conn.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
        finally:
            if 'cur' in locals():
                cur.close()
            release_db_connection(conn, db_pool)
            
    except Exception as e:
        logger.error(f"Error in delete_complaint: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# Add this helper function to debug data fetching
@app.route('/api/debug_complaints', methods=['GET'])
def debug_complaints():
    """Debug endpoint to check what data is being fetched"""
    try:
        conn = get_rasa_connection()
        if not conn:
            return jsonify({'error': 'Database connection failed'}), 500
        
        cur = conn.cursor()
        
        debug_info = {
            'power_outage_reports': [],
            'meter_concern': [],
            'agent_queue': []
        }
        
        # Check power_outage_reports
        cur.execute("""
            SELECT report_id, full_name, address, status, priority, timestamp, hidden
            FROM power_outage_reports
            ORDER BY timestamp DESC
            LIMIT 10
        """)
        debug_info['power_outage_reports'] = [
            {
                'id': row[0],
                'name': row[1],
                'address': row[2],
                'status': row[3],
                'priority': row[4],
                'timestamp': str(row[5]),
                'hidden': row[6]
            }
            for row in cur.fetchall()
        ]
        
        # Check meter_concern
        cur.execute("""
            SELECT id, name, address, status, priority, timestamp, hidden
            FROM meter_concern
            ORDER BY timestamp DESC
            LIMIT 10
        """)
        debug_info['meter_concern'] = [
            {
                'id': row[0],
                'name': row[1],
                'address': row[2],
                'status': row[3],
                'priority': row[4],
                'timestamp': str(row[5]),
                'hidden': row[6]
            }
            for row in cur.fetchall()
        ]
        
        # Check agent_queue
        cur.execute("""
            SELECT id, full_name, concern, status, priority, timestamp, hidden, resumed
            FROM agent_queue
            ORDER BY timestamp DESC
            LIMIT 10
        """)
        debug_info['agent_queue'] = [
            {
                'id': row[0],
                'name': row[1],
                'concern': row[2],
                'status': row[3],
                'priority': row[4],
                'timestamp': str(row[5]),
                'hidden': row[6],
                'resumed': row[7]
            }
            for row in cur.fetchall()
        ]
        
        cur.close()
        release_db_connection(conn, db_pool)
        
        return jsonify({
            'success': True,
            'debug_info': debug_info,
            'total_counts': {
                'power_outage': len(debug_info['power_outage_reports']),
                'meter_concern': len(debug_info['meter_concern']),
                'agent_queue': len(debug_info['agent_queue'])
            }
        })
        
    except Exception as e:
        logger.error(f"Debug endpoint error: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    



@app.route('/update_status/<table>/<int:record_id>', methods=['POST'])
def update_status(table, record_id):
    """Update the status of a complaint"""
    try:
        data = request.get_json()
        new_status = data.get('status', '').strip().upper()
        
        if not new_status:
            logger.warning("Status missing in update_status request")
            return jsonify({'success': False, 'error': 'Status is required'})
        
        # Validate status value
        valid_statuses = ['NEW', 'ASSIGNED', 'IN_PROGRESS', 'RESOLVED']
        if new_status not in valid_statuses:
            logger.warning(f"Invalid status: {new_status}")
            return jsonify({'success': False, 'error': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'})
        
        conn = get_rasa_connection()
        if not conn:
            logger.error("Database connection failed for update_status")
            return jsonify({'success': False, 'error': 'Database connection failed'})
        
        try:
            cur = conn.cursor()
            
            # Ensure the status column exists
            ensure_columns(conn, table)
            
            # Get the primary key column name
            pk_column = 'report_id' if table == 'power_outage_reports' else 'id'
            
            # ============================================
            # 🔴 ADD THIS: GET OLD STATUS FIRST
            # ============================================
            cur.execute(f"SELECT status FROM {table} WHERE {pk_column} = %s", (record_id,))
            result = cur.fetchone()
            old_status = result[0] if result else 'NEW'
            # ============================================
            
            # Check if record exists
            if not result:
                logger.warning(f"Record not found for status update: {table}/{record_id}")
                return jsonify({'success': False, 'error': 'Record not found'})
            
            # Update the status
            cur.execute(
                f"UPDATE {table} SET status = %s WHERE {pk_column} = %s",
                (new_status, record_id)
            )
            
            if cur.rowcount == 0:
                logger.warning(f"No rows updated for {table}/{record_id}")
                conn.rollback()
                return jsonify({'success': False, 'error': 'Failed to update status'})
            
            # Commit the changes
            conn.commit()
            
            # ============================================
            # 🔴 ADD THIS: BROADCAST STATUS CHANGE
            # ============================================
            broadcast_status_update(table, record_id, old_status, new_status)
            # ============================================
            
            logger.info(f"Status updated to {new_status} for {table}/{record_id}")
            
            return jsonify({
                'success': True,
                'message': f'Status updated to {new_status}',
                'new_status': new_status,
                'table': table,
                'record_id': record_id
            })
            
        except Exception as e:
            logger.error(f"Error updating status: {e}")
            traceback.print_exc()
            conn.rollback()
            return jsonify({'success': False, 'error': str(e)})
        finally:
            if 'cur' in locals():
                cur.close()
            release_db_connection(conn, db_pool)
            
    except Exception as e:
        logger.error(f"Error in update_status: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/hide_complaint/<table>/<int:record_id>', methods=['POST'])
def hide_complaint(table, record_id):
    """Hide a complaint (soft delete - can be recovered with Show Hidden checkbox)"""
    conn = get_rasa_connection()
    if not conn:
        logger.error("Database connection failed for hide_complaint")
        return jsonify({'success': False, 'error': 'Database connection failed'}), 500
    
    try:
        cur = conn.cursor()
        
        # Verify table exists
        cur.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_schema = 'public' AND table_name = %s
            )
        """, (table,))
        table_exists = cur.fetchone()[0]
        
        if not table_exists:
            logger.warning(f"Table does not exist: {table}")
            return jsonify({'success': False, 'error': f'Table {table} does not exist'}), 400
        
        # Get primary key column name
        pk_column = 'report_id' if table == 'power_outage_reports' else 'id'
        
        # Check if record exists
        cur.execute(f"SELECT {pk_column} FROM {table} WHERE {pk_column} = %s", (record_id,))
        if not cur.fetchone():
            logger.warning(f"Complaint not found: {table}/{record_id}")
            return jsonify({'success': False, 'error': 'Complaint not found'}), 404
        
        # Set hidden = TRUE (soft delete)
        cur.execute(f"UPDATE {table} SET hidden = TRUE WHERE {pk_column} = %s", (record_id,))
        conn.commit()
        
        logger.info(f"Complaint hidden: {table}/{record_id}")
        return jsonify({
            'success': True, 
            'message': 'Complaint hidden successfully (can be shown again with Show Hidden checkbox)'
        })
        
    except Exception as e:
        logger.error(f"Error hiding complaint: {e}")
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

def background_stats_broadcaster():
    """Background task to periodically broadcast stats"""
    while True:
        try:
            socketio.sleep(30)  # Every 30 seconds
            stats = get_current_stats()
            if stats:
                socketio.emit('stats_update', stats)
                logger.debug("📊 Periodic stats broadcast sent")
        except Exception as e:
            logger.error(f"Error in background broadcaster: {e}")
            
            # Start background task
socketio.start_background_task(background_stats_broadcaster)
        
if __name__ == '__main__':
    # Initialize database tables on startup
    logger.info("=" * 60)
    logger.info("Initializing database tables...")
    logger.info("=" * 60)
    
    try:
        # Existing initializations
        ensure_location_columns()
        ensure_issue_type_column()
        ensure_timestamp_column()
        create_indexes()
        ensure_converted_table_joblist()
        
        # Run incident details migration
        logger.info("🔧 Running incident details migration...")
        migrate_incident_details_columns()
        
        # Initialize PostGIS geometry support
        logger.info("🗺️ Initializing PostGIS geometry support...")
        initialize_postgis()
        
        logger.info("✓ Database initialization complete")
    except Exception as e:
        logger.error(f"✗ Database initialization failed: {e}")
        logger.warning("Some features may not work correctly")
    
    # Get configuration
    port = int(os.getenv('PORT', 5000))
    debug_mode = os.getenv('FLASK_DEBUG', 'True').lower() == 'true'
    host = os.getenv('FLASK_HOST', '0.0.0.0',)
    
    # Display startup information
    logger.info("=" * 60)
    logger.info("🚀 Starting ILECO-1 Chatbot Dashboard")
    logger.info("=" * 60)
    logger.info(f"Server Host: {host}")
    logger.info(f"Server Port: {port}")
    logger.info(f"Debug Mode: {debug_mode}")
    logger.info(f"Access URL: http://localhost:{port}")
    logger.info(f"Report Form: http://localhost:{port}/report")
    logger.info(f"Dashboard: http://localhost:{port}/")
    logger.info("=" * 60)
    logger.info("✅ Incident Details Feature: ENABLED")
    logger.info("=" * 60)
    
    # ============================================
    # 🔴 REPLACE app.run() WITH socketio.run()
    # ============================================
    logger.info("=" * 60)
    logger.info("🔴 Real-Time WebSocket Features ENABLED")
    logger.info("   - Live complaint counter")
    logger.info("   - New complaint notifications")
    logger.info("   - Real-time status updates")
    logger.info("   - Active users indicator")
    logger.info("=" * 60)
    
    # Start Flask application with SocketIO
    try:
        socketio.run(
            app,
            host=host,
            port=port,
            debug=debug_mode,
            use_reloader=debug_mode,
            allow_unsafe_werkzeug=True
        )
    except Exception as e:
        logger.critical(f"Failed to start Flask-SocketIO server: {e}")
        exit(1)
# ✅ Add this helper function to verify status updates
def verify_status_update(table, record_id):
    """Verify that a status was actually updated in the database"""
    conn = get_rasa_connection()
    if not conn:
        return None
    
    try:
        cur = conn.cursor()
        pk_column = 'report_id' if table == 'power_outage_reports' else 'id'
        cur.execute(f"SELECT status FROM {table} WHERE {pk_column} = %s", (record_id,))
        result = cur.fetchone()
        return result[0] if result else None
    except Exception as e:
        logger.error(f"Error verifying status: {e}")
        return None
    finally:
        if 'cur' in locals():
            cur.close()
        release_db_connection(conn, db_pool)

# ✅ Add status verification endpoint
@app.route('/api/verify_status/<table>/<int:record_id>', methods=['GET'])
def api_verify_status(table, record_id):
    """API endpoint to verify current status of a complaint"""
    try:
        status = verify_status_update(table, record_id)
        if status is None:
            return jsonify({
                'success': False,
                'error': 'Record not found or error occurred'
            }), 404
        
        return jsonify({
            'success': True,
            'status': status,
            'table': table,
            'record_id': record_id
        })
    except Exception as e:
        logger.error(f"Error in verify_status API: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ✅ Improved ensure_columns function with better error handling
def ensure_columns(conn, table):
    """Ensure status, hidden, and resumed columns exist"""
    try:
        cur = conn.cursor()
        
        # Ensure status column
        cur.execute(f"""
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name = '{table}' AND column_name = 'status'
                ) THEN
                    ALTER TABLE {table} ADD COLUMN status VARCHAR(50) DEFAULT 'NEW';
                    RAISE NOTICE 'Added status column to %', '{table}';
                END IF;
            END $$;
        """)
        
        # Ensure hidden column
        cur.execute(f"""
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name = '{table}' AND column_name = 'hidden'
                ) THEN
                    ALTER TABLE {table} ADD COLUMN hidden BOOLEAN DEFAULT FALSE;
                    RAISE NOTICE 'Added hidden column to %', '{table}';
                END IF;
            END $$;
        """)
        
        # For agent_queue, ensure resumed and job_order_id columns
        if table == 'agent_queue':
            cur.execute(f"""
                DO $$
                BEGIN
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name = 'agent_queue' AND column_name = 'resumed'
                    ) THEN
                        ALTER TABLE agent_queue ADD COLUMN resumed BOOLEAN DEFAULT FALSE;
                        RAISE NOTICE 'Added resumed column to agent_queue';
                    END IF;
                    
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name = 'agent_queue' AND column_name = 'job_order_id'
                    ) THEN
                        ALTER TABLE agent_queue ADD COLUMN job_order_id VARCHAR(50);
                        RAISE NOTICE 'Added job_order_id column to agent_queue';
                    END IF;
                END $$;
            """)
        
        conn.commit()
        logger.info(f"✅ Ensured columns for table {table}")
        return True
        
    except Exception as e:
        logger.error(f"❌ Error ensuring columns for {table}: {e}")
        traceback.print_exc()
        conn.rollback()
        return False
    finally:
        if 'cur' in locals():
            cur.close()

# ✅ Add logging middleware to track all status updates
@app.before_request
def log_request():
    if request.method in ['POST', 'PUT', 'PATCH'] and 'status' in request.path:
        logger.info(f"📝 Status update request: {request.method} {request.path}")
        logger.info(f"   Data: {request.get_json() if request.is_json else 'N/A'}")

@app.after_request
def log_response(response):
    if request.method in ['POST', 'PUT', 'PATCH'] and 'status' in request.path:
        logger.info(f"✅ Response: {response.status_code}")
        if response.is_json:
            logger.info(f"   Response data: {response.get_json()}")
    return response